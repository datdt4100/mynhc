import os
import re
import json
import io
import time
import secrets
import string
import threading
import unicodedata
from functools import wraps
from datetime import datetime, timezone, timedelta

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, jsonify, send_file, flash, Response
)
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import (
    create_engine, MetaData, Table, Column, Integer, Text,
    insert, select, update, delete, and_, or_, func, text
)
import openpyxl
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Tổ hợp môn học — cấu hình của trường
# ---------------------------------------------------------------------------

# Môn bắt buộc (tất cả học sinh đều phải học, không nằm trong tổ hợp)

# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-me")

# ── Concurrency & real-time ──
_reg_lock    = threading.Lock()   # serialise class registration writes
_change_ts   = [0.0]              # bumped on any schedule/class change
_change_lock = threading.Lock()

def _bump(event_type="class", grade=None):
    """Increment change timestamp and payload for SSE clients."""
    with _change_lock:
        _change_ts[0] = time.time()
        _change_ts.append({"type": event_type, "grade": grade, "ts": _change_ts[0]})
        if len(_change_ts) > 2:          # keep only latest payload
            _change_ts[1:] = [_change_ts[-1]]

DATABASE_URL = os.environ.get("DATABASE_URL", "sqlite:///classreg.db")
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

_engine_kwargs = {"future": True, "pool_pre_ping": True}
if not DATABASE_URL.startswith("sqlite"):
    _engine_kwargs.update({
        "pool_size": 5,
        "max_overflow": 2,
        "pool_recycle": 300,
        "pool_timeout": 30,
        "connect_args": {"sslmode": "require"},
    })
engine = create_engine(DATABASE_URL, **_engine_kwargs)
metadata = MetaData()

# ---------------------------------------------------------------------------
# Table definitions
# ---------------------------------------------------------------------------

teachers = Table(
    "teachers", metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("full_name", Text, nullable=False),
    Column("cccd", Text, unique=True, nullable=False),
    Column("gender", Text, nullable=False),           # Nam / Nữ
    Column("subject_group", Text, nullable=False),    # Tổ bộ môn
    Column("email", Text, nullable=True),
    Column("password_hash", Text, nullable=True),
    Column("is_first_login", Integer, default=1),
    Column("must_change_password", Integer, default=0),
    Column("activated_at", Text, nullable=True),
)

students = Table(
    "students", metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("full_name", Text, nullable=False),
    Column("cccd", Text, unique=True, nullable=False),
    Column("class_name", Text, nullable=False),
    Column("grade", Integer, nullable=False),         # 10 / 11 / 12
    Column("email", Text, nullable=True),
    Column("password_hash", Text, nullable=True),
    Column("is_first_login", Integer, default=1),
    Column("must_change_password", Integer, default=0),
    Column("activated_at", Text, nullable=True),
    Column("last_seen_at", Text, nullable=True),
)

classes = Table(
    "classes", metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("teacher_id", Integer, nullable=False),
    Column("grade", Integer, nullable=False),
    Column("duration", Integer, nullable=False),      # 1-4 tiết
    Column("day_of_week", Integer, nullable=False),   # 2-7
    Column("session_type", Text, nullable=False),     # morning / afternoon
    Column("start_session", Integer, nullable=False), # 1-4
    Column("subject", Text, nullable=True),
    Column("location", Text, nullable=True),
    Column("max_capacity", Integer, nullable=True),
    Column("extra_data", Text, nullable=True),        # JSON blob
    Column("is_published", Integer, default=0),
    Column("created_at", Text, default="CURRENT_TIMESTAMP"),
)

enrollments = Table(
    "enrollments", metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("student_id", Integer, nullable=False),
    Column("class_id", Integer, nullable=False),
    Column("enrolled_at", Text, default="CURRENT_TIMESTAMP"),
)

rooms = Table(
    "rooms", metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("name", Text, unique=True, nullable=False),
)

room_external_busy = Table(
    "room_external_busy", metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("room_name", Text, nullable=False),
    Column("day_of_week", Integer, nullable=False),   # 2-7
    Column("session_type", Text, nullable=False),     # morning / afternoon
    Column("tiet", Integer, nullable=False),          # 1-4
)

room_grade_slots = Table(
    "room_grade_slots", metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("room_name", Text, nullable=False),
    Column("day_of_week", Integer, nullable=False),
    Column("session_type", Text, nullable=False),
    Column("tiet", Integer, nullable=False),
    Column("available_grades", Text, nullable=False),  # "10", "11", "10,11"
)

settings_table = Table(
    "settings", metadata,
    Column("key", Text, primary_key=True),
    Column("value", Text, nullable=True),
)

operators = Table(
    "operators", metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("full_name", Text, nullable=False),
    Column("login_code", Text, unique=True, nullable=False),  # Mã đăng nhập (như CCCD)
    Column("password_hash", Text, nullable=False),
)

# ---------------------------------------------------------------------------
# DB init
# ---------------------------------------------------------------------------

def init_db():
    metadata.create_all(engine)
    with engine.connect() as conn:
        # Migrate: add columns if not exists (for existing DBs)
        for stmt in [
            "ALTER TABLE teachers ADD COLUMN must_change_password INTEGER DEFAULT 0",
            "ALTER TABLE students ADD COLUMN must_change_password INTEGER DEFAULT 0",
            "ALTER TABLE students ADD COLUMN email TEXT",
            "ALTER TABLE teachers ADD COLUMN activated_at TEXT",
            "ALTER TABLE students ADD COLUMN activated_at TEXT",
            "ALTER TABLE students ADD COLUMN last_seen_at TEXT",
        ]:
            try:
                conn.execute(text(stmt))
                conn.commit()
            except Exception:
                conn.rollback()

        # Migrate: update old default admin password
        try:
            conn.execute(text(
                "UPDATE settings SET value = 'Admin@123' WHERE key = 'admin_password' AND value = 'admin123'"
            ))
            conn.commit()
        except Exception:
            conn.rollback()

        # Migrate: create rooms table if not exists (for existing DBs)
        try:
            conn.execute(text(
                "CREATE TABLE IF NOT EXISTS rooms (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL)"
            ))
            conn.commit()
        except Exception:
            conn.rollback()

        # Seed default settings
        for key, value in [("admin_password", "Admin@123"),
                           ("teacher_reg_open", "0"),
                           ("student_reg_open", "0"),
                           ("maintenance_mode", "0"),
                           ("schedule_constraint", "1")]:
            try:
                conn.execute(
                    insert(settings_table).values(key=key, value=value)
                )
                conn.commit()
            except Exception:
                conn.rollback()

        # Seed sample data on fresh databases
        try:
            _seed_sample_data(conn)
        except Exception:
            conn.rollback()


def _seed_sample_data(conn):
    count = conn.execute(text("SELECT COUNT(*) FROM teachers")).scalar()
    if count and count > 0:
        return

    HASH_A = ("scrypt:32768:8:1$zuZMrDLImUA1908J$0e2304952a47da5e9cd0e6dbd6bff34ad931c32268b701eea265f9217710fda94349e12b374b950c0ecdce9b0e218644005748243eb97127dda1414bbbd0cb17")
    HASH_B = ("scrypt:32768:8:1$giVkCeMZteCUbYcn$dc3df2b2d7583f65eaa348b7ede6f82511ff52f262fb759024ccfa5f98ec33898926a65db03039df42faf73d034463b97b30f00e13bdbf0898f62d6abc4e530c")
    HASH_C = ("scrypt:32768:8:1$PIxoXbtxyA8eFs4W$e29a01ff956f6ba958858c770e1297b00315561b413a670eec8fe5ac33b72b251065ca9e3b6a4c632ca1f6dd00a45b1b3173c358bdfd4235634f20088073d44d")

    teacher_rows = [
        # id, full_name, cccd, gender, subject_group, email, pw_hash, is_first_login, must_change_pw
        (1,  "Nguyễn Thị Kim Oanh",      "079301000001", "Nữ",  "Hóa học",    "ntko1486@gmail.com",              HASH_A, 0, 0),
        (2,  "Nguyễn Phi Công",           "079301000002", "Nam", "Tiếng Anh",  "mrphicongnguyen@gmail.com",       HASH_A, 0, 0),
        (3,  "Lê Thị Hà",                 "079301000003", "Nữ",  "Hóa học",    "hahoanhc03@gmail.com",            HASH_A, 0, 0),
        (4,  "Cao Thị Hà",                "079301000004", "Nữ",  "Tiếng Anh",  "caothiha779@gmail.com",           HASH_A, 0, 0),
        (5,  "Hoàng Thị Thu Hằng",        "079301000005", "Nữ",  "Tiếng Anh",  "hoanghangnhc@gmail.com",          HASH_A, 0, 0),
        (6,  "Lê Thanh Nhã",              "079301000006", "Nữ",  "Toán",       "thanhnha4872@gmail.com",          HASH_A, 0, 0),
        (7,  "Nguyễn Minh Hiếu",          "079301000007", "Nam", "Ngữ Văn",    "nguyenhieu2312@gmail.com",        HASH_A, 0, 0),
        (8,  "Nguyễn Thị Hoài Dung",      "079301000008", "Nữ",  "Tiếng Anh",  "nguyendungnhc@gmail.com",         HASH_A, 0, 0),
        (9,  "Lâm Thị Hồng Loan",         "079301000009", "Nữ",  "Tiếng Anh",  "lamthihongloan2903@gmail.com",    HASH_A, 0, 0),
        (10, "Tạ Thanh Thúy",             "079301000010", "Nữ",  "Tiếng Anh",  "thuythuya04@gmail.com",           HASH_A, 0, 0),
        (11, "Lê Đông Hải",               "079301000011", "Nam", "Vật lý",     "ledonghai22091984@gmail.com",     HASH_A, 0, 0),
        (12, "Nguyễn Thị Kim Tuyền",      "079301000012", "Nữ",  "Toán",       "ntktuyen2008@gmail.com",          HASH_A, 0, 0),
        (13, "Vũ Thị Lý",                 "079301000013", "Nữ",  "Toán",       "vuthily1978@gmail.com",           HASH_A, 0, 0),
        (14, "Chu Thị Phương",            "079301000014", "Nữ",  "Ngữ Văn",    "phuongchu17@gmail.com",           HASH_A, 0, 0),
        (15, "Trần Quang Minh",           "079301000015", "Nam", "Hóa học",    "quangminhnhc@gmail.com",          HASH_A, 0, 0),
        (16, "Nguyễn Thị Mỹ Hạnh",       "079301000016", "Nữ",  "Vật lý",     "bhanhtk20@gmail.com",             HASH_A, 0, 0),
        (17, "Nguyễn Thanh Tú",           "079301000017", "Nam", "Vật lý",     "nguyenthanhtu66666666@gmail.com", HASH_A, 0, 0),
        (18, "Lê Thị Trúc Lâm",          "079301000018", "Nữ",  "Vật lý",     "lethitruclam38086@gmail.com",     HASH_A, 0, 0),
        (19, "Quách Huỳnh Hạnh",          "079301000019", "Nữ",  "Toán",       "huynhhanhquach@gmail.com",        HASH_A, 0, 0),
        (20, "Hứa Thị Hạ Phương",         "079301000020", "Nữ",  "Toán",       "phuonghua1587@gmail.com",         HASH_A, 0, 0),
        (21, "Nguyễn Đoàn Thùy Ngân",    "079301000021", "Nữ",  "Ngữ Văn",    "ngannhc2710@gmail.com",           HASH_A, 0, 0),
    ]
    for r in teacher_rows:
        conn.execute(text(
            "INSERT INTO teachers (id,full_name,cccd,gender,subject_group,email,password_hash,is_first_login,must_change_password) "
            "VALUES (:id,:fn,:cccd,:g,:sg,:em,:ph,:ifl,:mcp)"
        ), {"id": r[0], "fn": r[1], "cccd": r[2], "g": r[3], "sg": r[4], "em": r[5], "ph": r[6], "ifl": r[7], "mcp": r[8]})

    student_rows = [
        (1,  "Tran Van Binh",          "079200000002", "10C01", 10, None,                           None,   1, 0),
        (2,  "Nguyễn Văn An",          "079302000001", "10C01", 10, "an.nguyenvan@nhc.edu.vn",      HASH_A, 0, 0),
        (3,  "Trần Thị Bích",          "079302000002", "10C01", 10, "bich.tranthi@nhc.edu.vn",      HASH_A, 0, 0),
        (4,  "Lê Hoàng Cường",         "079302000003", "10C02", 10, None,                           None,   1, 0),
        (5,  "Phạm Ngọc Diệu",         "079302000004", "10C02", 10, "dieu.phamngoc@nhc.edu.vn",     HASH_A, 0, 0),
        (6,  "Võ Minh Đức",            "079302000005", "10C03", 10, None,                           None,   1, 0),
        (7,  "Nguyễn Hữu Tài",         "079311000001", "11A01", 11, "tai.nguyenhuu@nhc.edu.vn",     HASH_C, 0, 0),
        (8,  "Trần Thị Minh Châu",     "079311000002", "11A01", 11, "chau.tranthiminh@nhc.edu.vn",  HASH_C, 0, 0),
        (9,  "Lê Văn Khoa",            "079311000003", "11A02", 11, None,                           None,   1, 0),
        (10, "Đặng Thị Thu Hà",        "079311000004", "11A02", 11, None,                           None,   1, 0),
        (11, "Phan Thị Ngọc Lan",      "079312000001", "12A01", 12, "lan.phanthingoc@nhc.edu.vn",   HASH_C, 0, 0),
        (12, "Bùi Văn Thắng",          "079312000002", "12A01", 12, "thang.buivan@nhc.edu.vn",      HASH_C, 0, 0),
        (13, "Vũ Thị Tuyết Nhi",       "079312000003", "12A02", 12, None,                           None,   1, 0),
        (14, "Hoàng Minh Long",         "079312000004", "12A02", 12, None,                           None,   1, 0),
    ]
    for r in student_rows:
        conn.execute(text(
            "INSERT INTO students (id,full_name,cccd,class_name,grade,email,password_hash,is_first_login,must_change_password) "
            "VALUES (:id,:fn,:cccd,:cn,:gr,:em,:ph,:ifl,:mcp)"
        ), {"id": r[0], "fn": r[1], "cccd": r[2], "cn": r[3], "gr": r[4], "em": r[5], "ph": r[6], "ifl": r[7], "mcp": r[8]})

    class_rows = [
        # id, teacher_id, grade, duration, day_of_week, session_type, start_session, subject, location, max_cap
        # ── Khối 10 (19 lớp — dữ liệu production) ────────────────────────────
        (1,  1,  10, 2, 2, "afternoon", 1, "Hóa học",   "Phòng 1 - Khu A- Tầng 1",   50),  # T2 C t1-2
        (2,  2,  10, 2, 2, "morning",   3, "Tiếng Anh", "Phòng 13 - Khu A- Tầng 3",  50),  # T2 S t3-4
        (3,  3,  10, 2, 3, "afternoon", 1, "Hóa học",   "Phòng 32 - Khu D - Tầng 1", 50),  # T3 C t1-2
        (4,  4,  10, 2, 3, "afternoon", 1, "Tiếng Anh", "Phòng 1 - Khu A- Tầng 1",   50),  # T3 C t1-2
        (5,  5,  10, 1, 3, "afternoon", 3, "Tiếng Anh", "Phòng 33 - Khu D - Tầng 1", 50),  # T3 C t3
        (6,  6,  10, 2, 3, "afternoon", 3, "Toán",      "Phòng 32 - Khu D - Tầng 1", 50),  # T3 C t3-4
        (7,  7,  10, 1, 3, "afternoon", 4, "Ngữ Văn",   "Phòng 36 - Khu D - Tầng 2", 50),  # T3 C t4
        (8,  5,  10, 1, 3, "afternoon", 4, "Tiếng Anh", "Phòng 33 - Khu D - Tầng 1", 50),  # T3 C t4
        (9,  8,  10, 2, 4, "afternoon", 1, "Tiếng Anh", "Phòng 16 - Khu B - Tầng 1", 50),  # T4 C t1-2
        (10, 9,  10, 2, 4, "afternoon", 1, "Tiếng Anh", "Phòng 17 - Khu B - Tầng 1", 50),  # T4 C t1-2
        (11, 10, 10, 2, 5, "afternoon", 1, "Tiếng Anh", "Phòng 34 - Khu D - Tầng 1", 50),  # T5 C t1-2
        (12, 11, 10, 2, 5, "afternoon", 1, "Vật lý",    "Phòng 36 - Khu D - Tầng 2", 50),  # T5 C t1-2
        (13, 12, 10, 2, 5, "afternoon", 1, "Toán",      "Phòng 33 - Khu D - Tầng 1", 50),  # T5 C t1-2
        (14, 13, 10, 2, 5, "afternoon", 1, "Toán",      "Phòng 38 - Khu D - Tầng 2", 50),  # T5 C t1-2
        (15, 14, 10, 1, 5, "afternoon", 1, "Ngữ Văn",   "Phòng 32 - Khu D - Tầng 1", 50),  # T5 C t1
        (16, 4,  10, 4, 5, "afternoon", 1, "Tiếng Anh", "Phòng 2 - Khu A- Tầng 1",   50),  # T5 C t1-4
        (17, 16, 10, 2, 5, "afternoon", 1, "Vật lý",    "Phòng 37 - Khu D - Tầng 2", 50),  # T5 C t1-2
        (18, 14, 10, 1, 5, "afternoon", 2, "Ngữ Văn",   "Phòng 32 - Khu D - Tầng 1", 50),  # T5 C t2
        (19, 15, 10, 2, 5, "afternoon", 3, "Hóa học",   "Phòng 3 - Khu A- Tầng 1",   50),  # T5 C t3-4

        # ── Khối 11 (12 lớp — dữ liệu production) ────────────────────────────
        (20, 17, 11, 2, 3, "afternoon", 1, "Vật lý",    "Phòng 17 - Khu B - Tầng 1", 50),  # T3 C t1-2
        (21, 18, 11, 1, 3, "afternoon", 1, "Vật lý",    "Phòng 18 - Khu B - Tầng 1", 50),  # T3 C t1
        (22, 7,  11, 2, 3, "afternoon", 1, "Ngữ Văn",   "Phòng 11 - Khu A- Tầng 3",  50),  # T3 C t1-2
        (23, 18, 11, 1, 3, "afternoon", 2, "Vật lý",    "Phòng 18 - Khu B - Tầng 1", 50),  # T3 C t2
        (24, 7,  11, 1, 3, "afternoon", 3, "Ngữ Văn",   "Phòng 20 - Khu B - Tầng 2", 50),  # T3 C t3
        (25, 19, 11, 2, 4, "afternoon", 1, "Toán",      "Phòng 1 - Khu A- Tầng 1",   50),  # T4 C t1-2
        (26, 20, 11, 1, 4, "afternoon", 1, "Toán",      "Phòng 21 - Khu B - Tầng 2", 50),  # T4 C t1
        (27, 20, 11, 1, 4, "afternoon", 2, "Toán",      "Phòng 21 - Khu B - Tầng 2", 50),  # T4 C t2
        (28, 21, 11, 1, 4, "afternoon", 2, "Ngữ Văn",   "Phòng 18 - Khu B - Tầng 1", 50),  # T4 C t2
        (29, 10, 11, 2, 4, "afternoon", 3, "Tiếng Anh", "Phòng 17 - Khu B - Tầng 1", 50),  # T4 C t3-4
        (30, 21, 11, 1, 4, "afternoon", 3, "Ngữ Văn",   "Phòng 18 - Khu B - Tầng 1", 50),  # T4 C t3
        (31, 21, 11, 1, 4, "afternoon", 4, "Ngữ Văn",   "Phòng 18 - Khu B - Tầng 1", 50),  # T4 C t4

        # ── Khối 12 (2 lớp — dữ liệu production) ─────────────────────────────
        (32, 7,  12, 2, 5, "afternoon", 3, "Ngữ Văn",   "Phòng 27 - Khu C - Tầng 1", 50),  # T5 C t3-4
        (33, 7,  12, 1, 6, "afternoon", 1, "Ngữ Văn",   "Phòng 4 - Khu A- Tầng 1",   50),  # T6 C t1
    ]
    for r in class_rows:
        conn.execute(text(
            "INSERT INTO classes (id,teacher_id,grade,duration,day_of_week,session_type,start_session,subject,location,max_capacity,is_published) "
            "VALUES (:id,:tid,:gr,:dur,:dow,:st,:ss,:subj,:loc,:cap,1)"
        ), {"id": r[0], "tid": r[1], "gr": r[2], "dur": r[3], "dow": r[4], "st": r[5], "ss": r[6], "subj": r[7], "loc": r[8], "cap": r[9]})

    enrollment_rows = []  # no pre-seeded enrollments — students choose from production classes
    for r in enrollment_rows:
        conn.execute(text(
            "INSERT INTO enrollments (id,student_id,class_id) VALUES (:id,:sid,:cid)"
        ), {"id": r[0], "sid": r[1], "cid": r[2]})

    # Seed rooms
    room_names = [
        "Phòng 1 - Khu A- Tầng 1",   "Phòng 2 - Khu A- Tầng 1",
        "Phòng 3 - Khu A- Tầng 1",   "Phòng 4 - Khu A- Tầng 1",
        "Phòng 5 - Khu A- Tầng 1",   "Phòng 6 - Khu A- Tầng 2",
        "Phòng 7 - Khu A- Tầng 2",   "Phòng 8 - Khu A- Tầng 2",
        "Phòng 9 - Khu A- Tầng 2",   "Phòng 10 - Khu A- Tầng 2",
        "Phòng 11 - Khu A- Tầng 3",  "Phòng 12 - Khu A- Tầng 3",
        "Phòng 13 - Khu A- Tầng 3",  "Phòng 14 - Khu A- Tầng 3",
        "Phòng 15 - Khu A- Tầng 3",  "Phòng 44 - Khu A - Tầng Trệt",
        "Hội Trường 1 - Khu A - Tầng Trệt",
        "Phòng 16 - Khu B - Tầng 1", "Phòng 17 - Khu B - Tầng 1",
        "Phòng 18 - Khu B - Tầng 1", "Phòng 19 - Khu B - Tầng 2",
        "Phòng 20 - Khu B - Tầng 2", "Phòng 21 - Khu B - Tầng 2",
        "Phòng 22 - Khu B - Tầng 3", "Phòng 23 - Khu B - Tầng 3",
        "Phòng 24 - Khu B - Tầng 3",
        "Phòng 25 - Khu C - Tầng 1", "Phòng 26 - Khu C - Tầng 1",
        "Phòng 27 - Khu C - Tầng 1", "Phòng 28 - Khu C - Tầng 3",
        "Phòng 29 - Khu C - Tầng 3", "Phòng 30 - Khu C - Tầng 3",
        "Phòng 31 - Khu C - Tầng 3",
        "Phòng 32 - Khu D - Tầng 1", "Phòng 33 - Khu D - Tầng 1",
        "Phòng 34 - Khu D - Tầng 1", "Phòng 35 - Khu D - Tầng 1",
        "Phòng 36 - Khu D - Tầng 2", "Phòng 37 - Khu D - Tầng 2",
        "Phòng 38 - Khu D - Tầng 2", "Phòng 39 - Khu D - Tầng 2",
        "Phòng 40 - Khu D - Tầng 3", "Phòng 41 - Khu D - Tầng 3",
        "Phòng 42 - Khu D - Tầng 3", "Phòng 43 - Khu D - Tầng 3",
        "Hội Trường 2 - Khu E - Tầng 2",
    ]
    for name in room_names:
        try:
            conn.execute(text("INSERT INTO rooms (name) VALUES (:n)"), {"n": name})
            conn.commit()
        except Exception:
            conn.rollback()

    # Reset PostgreSQL sequences after explicit-ID inserts
    if not DATABASE_URL.startswith("sqlite"):
        for tbl in ("teachers", "students", "classes", "enrollments"):
            conn.execute(text(
                f"SELECT setval(pg_get_serial_sequence('{tbl}', 'id'), MAX(id)) FROM {tbl}"
            ))

    conn.commit()


with app.app_context():
    init_db()

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _room_sort_key(name: str):
    if re.search(r'[Hh]ội\s*[Tt]rường', name):
        num = re.search(r'(\d+)', name)
        return ('0', int(num.group(1)) if num else 0)
    khu = re.search(r'[Kk]hu\s+([A-Za-z])', name)
    num = re.search(r'(\d+)', name)
    return (khu.group(1).upper() if khu else 'Z', int(num.group(1)) if num else 9999)

_VN_TZ = timezone(timedelta(hours=7))

def now_vn() -> str:
    return datetime.now(_VN_TZ).strftime("%Y-%m-%d %H:%M:%S")

# Allowed special chars (excludes ' " ` ; \ which are DB-dangerous)
_PW_ALLOWED_SPECIALS = r"!@#$%^&*()\-_+=\[\]{}|<>,.?/~"
_PW_PATTERN = re.compile(r'^[A-Za-z0-9' + _PW_ALLOWED_SPECIALS + r']+$')


def validate_password(pw: str):
    """Normalize and validate password. Returns (ok, error_msg)."""
    pw = unicodedata.normalize('NFKC', pw).strip()
    if len(pw) < 8:
        return False, "Mật khẩu phải có ít nhất 8 ký tự."
    if not _PW_PATTERN.match(pw):
        return False, (
            "Mật khẩu chỉ được chứa chữ hoa, chữ thường, chữ số "
            "và ký tự đặc biệt hợp lệ (!@#$%^&*()-_+=[]{}|<>,.?/~). "
            "Không dùng dấu nháy (', \"), dấu chấm phẩy (;) hoặc dấu gạch chéo (\\)."
        )
    return True, None


def normalize_password(pw: str) -> str:
    return unicodedata.normalize('NFKC', pw).strip()


def generate_temp_password(length: int = 12) -> str:
    alphabet = string.ascii_letters + string.digits
    return ''.join(secrets.choice(alphabet) for _ in range(length))


def get_setting(key, default=None):
    with engine.connect() as conn:
        row = conn.execute(
            select(settings_table.c.value).where(settings_table.c.key == key)
        ).fetchone()
    return row[0] if row else default


def set_setting(key, value):
    with engine.connect() as conn:
        existing = conn.execute(
            select(settings_table.c.key).where(settings_table.c.key == key)
        ).fetchone()
        if existing:
            conn.execute(
                update(settings_table).where(settings_table.c.key == key).values(value=value)
            )
        else:
            conn.execute(insert(settings_table).values(key=key, value=value))
        conn.commit()


def _is_admin():
    return session.get("is_admin") is True


def day_name(n):
    mapping = {2: "Thứ 2", 3: "Thứ 3", 4: "Thứ 4",
               5: "Thứ 5", 6: "Thứ 6", 7: "Thứ 7"}
    return mapping.get(int(n), str(n))


def session_label(session_type, start, duration):
    label = "Sáng" if session_type == "morning" else "Chiều"
    end = int(start) + int(duration) - 1
    if int(duration) == 1:
        return f"{label}, tiết {start}"
    return f"{label}, tiết {start}-{end}"


def _time_conflict(student_id, new_class):
    """Return the conflicting class row if any, else None."""
    with engine.connect() as conn:
        enrolled = conn.execute(
            select(classes).join(
                enrollments, classes.c.id == enrollments.c.class_id
            ).where(enrollments.c.student_id == student_id)
        ).fetchall()

    new_day = int(new_class["day_of_week"])
    new_session = new_class["session_type"]
    new_start = int(new_class["start_session"])
    new_end = new_start + int(new_class["duration"]) - 1

    for c in enrolled:
        if int(c.day_of_week) != new_day:
            continue
        if c.session_type != new_session:
            continue
        c_start = int(c.start_session)
        c_end = c_start + int(c.duration) - 1
        # Overlap check
        if new_start <= c_end and c_start <= new_end:
            return c
    return None


def _norm_subj(s):
    return s.strip() if s else s


def _slots_overlap_dict(a, b):
    if int(a["day_of_week"]) != int(b["day_of_week"]):
        return False
    if a["session_type"] != b["session_type"]:
        return False
    a_end = int(a["start_session"]) + int(a["duration"]) - 1
    b_end = int(b["start_session"]) + int(b["duration"]) - 1
    return int(a["start_session"]) <= b_end and int(b["start_session"]) <= a_end


def _build_by_subj(grade, extra_class=None):
    with engine.connect() as conn:
        rows = conn.execute(
            select(classes.c.day_of_week, classes.c.session_type,
                   classes.c.start_session, classes.c.duration, classes.c.subject)
            .where(classes.c.grade == grade)
        ).fetchall()
    by_subj = {}
    for r in rows:
        d = {"day_of_week": r.day_of_week, "session_type": r.session_type,
             "start_session": r.start_session, "duration": r.duration}
        by_subj.setdefault(_norm_subj(r.subject), []).append(d)
    if extra_class:
        d = {k: extra_class[k] for k in ("day_of_week", "session_type", "start_session", "duration")}
        by_subj.setdefault(_norm_subj(extra_class.get("subject", "")), []).append(d)
    return by_subj


def _backtrack(by_subj, subjects, idx, selected, count, cap):
    if count[0] >= cap:
        return
    if idx == len(subjects):
        count[0] += 1
        return
    for cls in by_subj[subjects[idx]]:
        if not any(_slots_overlap_dict(cls, s) for s in selected):
            selected.append(cls)
            _backtrack(by_subj, subjects, idx + 1, selected, count, cap)
            selected.pop()


def count_valid_combos(grade, extra_class=None, cap=100):
    """Total valid combinations for grade after adding extra_class."""
    by_subj = _build_by_subj(grade, extra_class)
    subjects = sorted(by_subj.keys())
    if not subjects:
        return 0
    count = [0]
    _backtrack(by_subj, subjects, 0, [], count, cap)
    return count[0]


def count_combos_including(grade, new_class, cap=100):
    """
    Đếm số cách HS có thể chọn mỗi môn 1 lớp (tất cả môn hiện có) mà không trùng giờ,
    trong đó bắt buộc chọn new_class. Trả về 0 nếu không có cách nào hợp lệ.
    """
    by_subj = _build_by_subj(grade, new_class)
    new_slot = {k: new_class[k] for k in ("day_of_week", "session_type",
                                           "start_session", "duration")}
    subj = _norm_subj(new_class.get("subject", ""))
    others = sorted(s for s in by_subj if s != subj)
    if not others:
        return 1
    count = [0]
    _backtrack(by_subj, others, 0, [new_slot], count, cap)
    return count[0]


def _class_impact(grade, cls_dict, cap=100):
    """Số cách hợp lệ khi lớp đã có trong DB (dùng cho badge admin)."""
    by_subj = _build_by_subj(grade)
    fix_slot = {k: cls_dict[k] for k in ("day_of_week", "session_type",
                                          "start_session", "duration")}
    subj = _norm_subj(cls_dict.get("subject", ""))
    others = sorted(s for s in by_subj if s != subj)
    if not others:
        return cap
    count = [0]
    _backtrack(by_subj, others, 0, [fix_slot], count, cap)
    return count[0]


def count_combos_max(grade, subject, cap=100):
    """Số combo tối đa có thể có giữa các môn khác (không tính môn subject).
    Đây là mức trần — slot nào cho actual == max thì không làm mất combo nào."""
    by_subj = _build_by_subj(grade)
    subj = _norm_subj(subject or "")
    others = sorted(s for s in by_subj if s != subj)
    if not others:
        return 1  # Không có môn khác → tối đa 1 tổ hợp (chính lớp mới)
    count = [0]
    _backtrack(by_subj, others, 0, [], count, cap)
    return count[0] or 1  # Đảm bảo không chia cho 0


def teacher_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("user_type") != "teacher":
            return redirect(url_for("login_page"))
        if get_setting("maintenance_mode", "0") == "1":
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated


def student_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("user_type") != "student":
            return redirect(url_for("login_page"))
        if get_setting("maintenance_mode", "0") == "1":
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not _is_admin():
            return redirect(url_for("admin_login_page"))
        return f(*args, **kwargs)
    return decorated


def operator_required(f):
    """Allow only operator accounts (not admin)."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("user_type") == "operator":
            return f(*args, **kwargs)
        return redirect(url_for("login_page"))
    return decorated

# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return redirect(url_for("login_page"))


@app.route("/stream/changes")
def stream_changes():
    """SSE: push a JSON event whenever the schedule or class list changes."""
    def gen():
        last = 0.0
        while True:
            ts = _change_ts[0]
            if ts != last:
                last = ts
                payload = _change_ts[1] if len(_change_ts) > 1 else {"type": "ping", "ts": ts}
                yield f"data: {json.dumps(payload)}\n\n"
            else:
                yield ": ping\n\n"
            time.sleep(2)
    return Response(
        gen(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/login")
def login_page():
    maintenance = get_setting("maintenance_mode", "0") == "1"
    return render_template("login.html", maintenance=maintenance)


@app.route("/login/step1", methods=["POST"])
def login_step1():
    if get_setting("maintenance_mode", "0") == "1":
        return jsonify(ok=False, error="Hệ thống đang bảo trì. Vui lòng quay lại sau.")
    data = request.get_json(force=True)
    full_name = (data.get("full_name") or "").strip()
    cccd = (data.get("cccd") or "").strip()

    if not full_name or not cccd:
        return jsonify(ok=False, error="Vui lòng nhập đầy đủ thông tin.")

    with engine.connect() as conn:
        teacher = conn.execute(
            select(teachers).where(
                and_(teachers.c.full_name == full_name, teachers.c.cccd == cccd)
            )
        ).fetchone()
        if teacher:
            title = "Cô" if teacher.gender == "Nữ" else "Thầy"
            return jsonify(
                ok=True,
                user_type="teacher",
                full_name=teacher.full_name,
                gender=teacher.gender,
                title=title,
                is_first_login=teacher.is_first_login,
            )

        student = conn.execute(
            select(students).where(
                and_(students.c.full_name == full_name, students.c.cccd == cccd)
            )
        ).fetchone()
        if student:
            return jsonify(
                ok=True,
                user_type="student",
                full_name=student.full_name,
                gender=None,
                title="Học sinh",
                is_first_login=student.is_first_login,
            )

        operator = conn.execute(
            select(operators).where(
                and_(operators.c.full_name == full_name, operators.c.login_code == cccd)
            )
        ).fetchone()
        if operator:
            return jsonify(
                ok=True,
                user_type="operator",
                full_name=operator.full_name,
                gender=None,
                title="Quản trị lớp học",
                is_first_login=0,
            )

    return jsonify(ok=False, error="Không tìm thấy tài khoản phù hợp.")


@app.route("/login/step2", methods=["POST"])
def login_step2():
    if get_setting("maintenance_mode", "0") == "1":
        return jsonify(ok=False, error="Hệ thống đang bảo trì. Vui lòng quay lại sau.")
    data = request.get_json(force=True)
    full_name = (data.get("full_name") or "").strip()
    cccd = (data.get("cccd") or "").strip()
    user_type = data.get("user_type")
    password = normalize_password(data.get("password") or "")
    email = (data.get("email") or "").strip()

    if user_type == "teacher":
        with engine.connect() as conn:
            teacher = conn.execute(
                select(teachers).where(
                    and_(teachers.c.full_name == full_name, teachers.c.cccd == cccd)
                )
            ).fetchone()

        if not teacher:
            return jsonify(ok=False, error="Không tìm thấy giáo viên.")

        if teacher.is_first_login:
            # First login: set email + password
            if not email:
                return jsonify(ok=False, error="Vui lòng nhập email.")
            ok, err = validate_password(password)
            if not ok:
                return jsonify(ok=False, error=err)
            pw_hash = generate_password_hash(password)
            with engine.connect() as conn:
                conn.execute(
                    update(teachers).where(teachers.c.id == teacher.id).values(
                        email=email,
                        password_hash=pw_hash,
                        is_first_login=0,
                        must_change_password=0,
                        activated_at=now_vn(),
                    )
                )
                conn.commit()
        else:
            if not teacher.password_hash:
                return jsonify(ok=False, error="Tài khoản chưa được thiết lập mật khẩu.")
            if not check_password_hash(teacher.password_hash, password):
                return jsonify(ok=False, error="Sai mật khẩu.")

        session.clear()
        session["user_type"] = "teacher"
        session["user_id"] = teacher.id
        session["full_name"] = teacher.full_name
        session["gender"] = teacher.gender
        session["subject_group"] = teacher.subject_group

        must_change = getattr(teacher, 'must_change_password', 0) or 0
        if must_change and not teacher.is_first_login:
            return jsonify(ok=True, must_change=True, redirect=url_for("change_password_page"))
        return jsonify(ok=True, redirect=url_for("teacher_dashboard"))

    elif user_type == "student":
        with engine.connect() as conn:
            student = conn.execute(
                select(students).where(
                    and_(students.c.full_name == full_name, students.c.cccd == cccd)
                )
            ).fetchone()

        if not student:
            return jsonify(ok=False, error="Không tìm thấy học sinh.")

        if student.is_first_login:
            # First login: set password (no email required for students)
            ok, err = validate_password(password)
            if not ok:
                return jsonify(ok=False, error=err)
            pw_hash = generate_password_hash(password)
            with engine.begin() as conn:
                conn.execute(
                    update(students).where(students.c.id == student.id).values(
                        password_hash=pw_hash,
                        is_first_login=0,
                        must_change_password=0,
                        activated_at=now_vn(),
                        last_seen_at=now_vn(),
                    )
                )
        else:
            if not student.password_hash:
                return jsonify(ok=False, error="Tài khoản chưa được thiết lập mật khẩu.")
            if not check_password_hash(student.password_hash, password):
                return jsonify(ok=False, error="Sai mật khẩu.")

        if not student.is_first_login:
            with engine.begin() as conn:
                conn.execute(
                    update(students).where(students.c.id == student.id).values(
                        last_seen_at=now_vn()
                    )
                )

        session.clear()
        session["user_type"] = "student"
        session["user_id"] = student.id
        session["full_name"] = student.full_name
        session["grade"] = student.grade

        must_change = getattr(student, 'must_change_password', 0) or 0
        if must_change and not student.is_first_login:
            return jsonify(ok=True, must_change=True, redirect=url_for("change_password_page"))
        return jsonify(ok=True, redirect=url_for("student_dashboard"))

    elif user_type == "operator":
        with engine.connect() as conn:
            op = conn.execute(
                select(operators).where(
                    and_(operators.c.full_name == full_name, operators.c.login_code == cccd)
                )
            ).fetchone()
        if not op:
            return jsonify(ok=False, error="Không tìm thấy tài khoản.")
        if not check_password_hash(op.password_hash, password):
            return jsonify(ok=False, error="Sai mật khẩu.")
        session.clear()
        session["user_type"] = "operator"
        session["operator_id"] = op.id
        session["full_name"] = op.full_name
        return jsonify(ok=True, redirect=url_for("op_room_detail"))

    return jsonify(ok=False, error="Loại tài khoản không hợp lệ.")


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return redirect(url_for("login_page"))


@app.route("/change-password", methods=["GET"])
def change_password_page():
    user_type = session.get("user_type")
    if user_type not in ("teacher", "student"):
        return redirect(url_for("login_page"))
    return render_template("change_password.html")


@app.route("/change-password", methods=["POST"])
def change_password_submit():
    user_type = session.get("user_type")
    user_id   = session.get("user_id")
    if user_type not in ("teacher", "student") or not user_id:
        return jsonify(ok=False, error="Phiên đăng nhập hết hạn.")

    data = request.get_json(force=True)
    new_pw  = normalize_password(data.get("new_password") or "")
    confirm = normalize_password(data.get("confirm_password") or "")

    ok, err = validate_password(new_pw)
    if not ok:
        return jsonify(ok=False, error=err)
    if new_pw != confirm:
        return jsonify(ok=False, error="Mật khẩu xác nhận không khớp.")

    pw_hash = generate_password_hash(new_pw)
    table = teachers if user_type == "teacher" else students
    with engine.begin() as conn:
        conn.execute(
            update(table).where(table.c.id == user_id).values(
                password_hash=pw_hash,
                must_change_password=0,
            )
        )

    redirect_url = url_for("teacher_dashboard") if user_type == "teacher" else url_for("student_dashboard")
    return jsonify(ok=True, redirect=redirect_url)

# ---------------------------------------------------------------------------
# Teacher routes
# ---------------------------------------------------------------------------

@app.route("/teacher")
@teacher_required
def teacher_dashboard():
    teacher_id = session["user_id"]
    with engine.connect() as conn:
        teacher_row = conn.execute(
            select(teachers).where(teachers.c.id == teacher_id)
        ).fetchone()

        teacher_classes = conn.execute(
            select(classes).where(classes.c.teacher_id == teacher_id)
            .order_by(classes.c.grade, classes.c.day_of_week, classes.c.session_type, classes.c.start_session)
        ).fetchall()

        enrollment_counts = {}
        for c in teacher_classes:
            cnt = conn.execute(
                select(func.count()).where(enrollments.c.class_id == c.id)
            ).scalar()
            enrollment_counts[c.id] = cnt

    # Build schedule grid for Outlook-style calendar
    schedule = {}
    for c in teacher_classes:
        k = f"{c.day_of_week}_{c.session_type}_{c.start_session}"
        schedule[k] = {"status": "start", "cls": c, "rowspan": c.duration}
        for t in range(c.start_session + 1, c.start_session + c.duration):
            schedule[f"{c.day_of_week}_{c.session_type}_{t}"] = {"status": "blocked"}

    teacher_reg_open = get_setting("teacher_reg_open", "0") == "1"
    maintenance = get_setting("maintenance_mode", "0") == "1"
    schedule_constraint = get_setting("schedule_constraint", "1") == "1"
    return render_template(
        "teacher/dashboard.html",
        teacher=teacher_row,
        classes=teacher_classes,
        enrollment_counts=enrollment_counts,
        schedule=schedule,
        teacher_reg_open=teacher_reg_open,
        maintenance=maintenance,
        schedule_constraint=schedule_constraint,
        day_name=day_name,
        session_label=session_label,
    )


@app.route("/teacher/classes", methods=["POST"])
@teacher_required
def teacher_register_class():
    if get_setting("teacher_reg_open", "0") != "1":
        return jsonify(ok=False, error="Chưa mở đăng ký lớp.")

    data = request.get_json(force=True)
    teacher_id = session["user_id"]

    try:
        grade = int(data["grade"])
        duration = int(data["duration"])
        day_of_week = int(data["day_of_week"])
        session_type = data["session_type"]
        start_session = int(data["start_session"])
    except (KeyError, ValueError, TypeError):
        return jsonify(ok=False, error="Dữ liệu không hợp lệ.")

    if grade not in (10, 11, 12):
        return jsonify(ok=False, error="Khối phải là 10, 11 hoặc 12.")
    if duration not in (2, 4):
        return jsonify(ok=False, error="Số tiết phải là 2 hoặc 4.")
    if not (2 <= day_of_week <= 7):
        return jsonify(ok=False, error="Thứ phải từ 2 đến 7.")
    if session_type not in ("morning", "afternoon"):
        return jsonify(ok=False, error="Buổi không hợp lệ.")
    if duration == 2 and start_session not in (1, 3):
        return jsonify(ok=False, error="2 tiết chỉ được bắt đầu từ tiết 1 (tiết 1–2) hoặc tiết 3 (tiết 3–4).")
    if duration == 4 and start_session != 1:
        return jsonify(ok=False, error="4 tiết phải bắt đầu từ tiết 1.")
    if start_session + duration - 1 > 4:
        return jsonify(ok=False, error="Tiết kết thúc vượt quá tiết 4.")

    subject  = session.get("subject_group") or (data.get("subject") or "").strip() or None
    end_session = start_session + duration - 1

    # Heavy combo check — only when admin has not disabled the constraint
    if subject and get_setting("schedule_constraint", "1") == "1":
        new_cls = {"subject": subject, "day_of_week": day_of_week,
                   "session_type": session_type, "start_session": start_session,
                   "duration": duration}
        combos_inc = count_combos_including(grade, new_cls)
        if combos_inc == 0:
            return jsonify(ok=False,
                error=f"Học sinh khối {grade} chọn lớp này sẽ không thể ghép đủ các môn khác "
                      f"(0 tổ hợp hợp lệ). Vui lòng chọn thứ/tiết khác.")

    # Acquire write lock — serialise concurrent registrations
    with _reg_lock:
        with engine.begin() as conn:
            # Check teacher schedule conflict (same teacher, overlapping slot)
            teacher_conflict = conn.execute(
                select(classes.c.id).where(and_(
                    classes.c.teacher_id == teacher_id,
                    classes.c.day_of_week == day_of_week,
                    classes.c.session_type == session_type,
                    classes.c.start_session <= end_session,
                    (classes.c.start_session + classes.c.duration - 1) >= start_session,
                ))
            ).first()
            if teacher_conflict:
                return jsonify(ok=False,
                    error="Bạn đã có lớp trong khung giờ này. Vui lòng chọn khung giờ khác.")

            # Verify at least 1 room is free for this grade + slot
            all_rooms_rows = conn.execute(select(rooms)).fetchall()
            if all_rooms_rows:
                booked_locs = {r.location for r in conn.execute(
                    select(classes.c.location).where(and_(
                        classes.c.day_of_week == day_of_week,
                        classes.c.session_type == session_type,
                        classes.c.location.isnot(None),
                        classes.c.start_session <= end_session,
                        (classes.c.start_session + classes.c.duration - 1) >= start_session,
                    ))
                ).fetchall() if r.location}
                ext_busy_locs = {r.room_name for r in conn.execute(
                    select(room_external_busy.c.room_name).where(and_(
                        room_external_busy.c.day_of_week == day_of_week,
                        room_external_busy.c.session_type == session_type,
                        room_external_busy.c.tiet >= start_session,
                        room_external_busy.c.tiet <= end_session,
                    ))
                ).fetchall()}
                grade_blocked_locs = set()
                for gr in conn.execute(
                    select(room_grade_slots.c.room_name, room_grade_slots.c.available_grades).where(and_(
                        room_grade_slots.c.day_of_week == day_of_week,
                        room_grade_slots.c.session_type == session_type,
                        room_grade_slots.c.tiet >= start_session,
                        room_grade_slots.c.tiet <= end_session,
                    ))
                ).fetchall():
                    allowed = [g.strip() for g in gr.available_grades.split(",")]
                    if str(grade) not in allowed:
                        grade_blocked_locs.add(gr.room_name)
                free_rooms = [r.name for r in all_rooms_rows
                              if r.name not in booked_locs
                              and r.name not in ext_busy_locs
                              and r.name not in grade_blocked_locs]
                if not free_rooms:
                    return jsonify(ok=False, error="Không còn phòng trống cho khung giờ này.")

            result = conn.execute(
                insert(classes).values(
                    teacher_id=teacher_id,
                    grade=grade,
                    duration=duration,
                    day_of_week=day_of_week,
                    session_type=session_type,
                    start_session=start_session,
                    subject=subject,
                    location=None,
                    max_capacity=50,
                    extra_data=None,
                    is_published=1,
                    created_at=now_vn(),
                )
            )
            class_id = result.inserted_primary_key[0]

    _bump(event_type="class", grade=grade)
    return jsonify(ok=True, class_id=class_id)


@app.route("/teacher/classes/<int:class_id>", methods=["DELETE"])
@teacher_required
def teacher_delete_class(class_id):
    if get_setting("teacher_reg_open", "0") != "1":
        return jsonify(ok=False, error="Đăng ký lớp đã đóng, không thể thực hiện thay đổi.")
    teacher_id = session["user_id"]
    with engine.connect() as conn:
        cls = conn.execute(
            select(classes).where(
                and_(classes.c.id == class_id, classes.c.teacher_id == teacher_id)
            )
        ).fetchone()
        if not cls:
            return jsonify(ok=False, error="Không tìm thấy lớp.")

        enroll_count = conn.execute(
            select(func.count()).where(enrollments.c.class_id == class_id)
        ).scalar()
        if enroll_count > 0:
            return jsonify(ok=False, error="Lớp đã có học sinh đăng ký, không thể xóa.")

        conn.execute(delete(classes).where(classes.c.id == class_id))
        conn.commit()

    return jsonify(ok=True)


@app.route("/teacher/classes/<int:class_id>/students")
@teacher_required
def teacher_class_students(class_id):
    teacher_id = session["user_id"]
    with engine.connect() as conn:
        cls = conn.execute(
            select(classes).where(
                and_(classes.c.id == class_id, classes.c.teacher_id == teacher_id)
            )
        ).fetchone()
        if not cls:
            return jsonify(ok=False, error="Không tìm thấy lớp.")

        enrolled = conn.execute(
            select(students, enrollments.c.enrolled_at)
            .join(enrollments, students.c.id == enrollments.c.student_id)
            .where(enrollments.c.class_id == class_id)
            .order_by(students.c.full_name)
        ).fetchall()

    return jsonify(ok=True, students=[
        {
            "id": s.id,
            "full_name": s.full_name,
            "class_name": s.class_name,
            "grade": s.grade,
            "enrolled_at": s.enrolled_at,
        }
        for s in enrolled
    ])

# ---------------------------------------------------------------------------
# Student routes
# ---------------------------------------------------------------------------

@app.route("/student")
@student_required
def student_dashboard():
    student_id = session["user_id"]
    student_grade = session.get("grade")

    with engine.connect() as conn:
        student_row = conn.execute(
            select(students).where(students.c.id == student_id)
        ).fetchone()

        published_classes = conn.execute(
            select(classes, teachers.c.full_name.label("teacher_name"),
                   teachers.c.subject_group)
            .join(teachers, classes.c.teacher_id == teachers.c.id)
            .where(and_(
                classes.c.is_published == 1,
                classes.c.grade == student_grade,
                classes.c.location.isnot(None),
                classes.c.location != '',
                classes.c.max_capacity.isnot(None),
            ))
            .order_by(classes.c.day_of_week, classes.c.session_type, classes.c.start_session)
        ).fetchall()

        enrollment_counts = {}
        for c in published_classes:
            cnt = conn.execute(
                select(func.count()).where(enrollments.c.class_id == c.id)
            ).scalar()
            enrollment_counts[c.id] = cnt

        my_enrollments = conn.execute(
            select(enrollments.c.class_id).where(enrollments.c.student_id == student_id)
        ).fetchall()
        my_class_ids = {row.class_id for row in my_enrollments}

        student_schedule = {}
        for c in published_classes:
            if c.id in my_class_ids:
                for i in range(c.duration):
                    tiet = c.start_session + i
                    key = f"{c.day_of_week}_{c.session_type}_{tiet}"
                    if i == 0:
                        student_schedule[key] = {"status": "start", "cls": c, "rowspan": c.duration}
                    else:
                        student_schedule[key] = {"status": "blocked"}

    student_reg_open = get_setting("student_reg_open", "0") == "1"
    return render_template(
        "student/dashboard.html",
        student=student_row,
        published_classes=published_classes,
        enrollment_counts=enrollment_counts,
        my_class_ids=my_class_ids,
        student_schedule=student_schedule,
        student_reg_open=student_reg_open,
        day_name=day_name,
        session_label=session_label,
    )


@app.route("/student/enroll", methods=["POST"])
@student_required
def student_enroll():
    if get_setting("student_reg_open", "0") != "1":
        return jsonify(ok=False, error="Chưa mở đăng ký.")

    data = request.get_json(force=True)
    student_id = session["user_id"]
    student_grade = session.get("grade")

    try:
        class_id = int(data["class_id"])
    except (KeyError, ValueError, TypeError):
        return jsonify(ok=False, error="Dữ liệu không hợp lệ.")

    with engine.connect() as conn:
        cls = conn.execute(
            select(classes).where(classes.c.id == class_id)
        ).fetchone()

        if not cls:
            return jsonify(ok=False, error="Không tìm thấy lớp.")
        if not cls.is_published:
            return jsonify(ok=False, error="Lớp chưa được mở đăng ký.")
        if cls.grade != student_grade:
            return jsonify(ok=False, error="Lớp không thuộc khối của bạn.")
        if not cls.location or not cls.max_capacity:
            return jsonify(ok=False, error="Lớp chưa cập nhật đầy đủ thông tin.")

        # Check already enrolled
        existing = conn.execute(
            select(enrollments).where(
                and_(enrollments.c.student_id == student_id,
                     enrollments.c.class_id == class_id)
            )
        ).fetchone()
        if existing:
            return jsonify(ok=False, error="Bạn đã đăng ký lớp này rồi.")

        # Check capacity
        if cls.max_capacity is not None:
            cnt = conn.execute(
                select(func.count()).where(enrollments.c.class_id == class_id)
            ).scalar()
            if cnt >= cls.max_capacity:
                return jsonify(ok=False, error="Lớp đã đầy.")

    # Check time conflict
    conflict = _time_conflict(student_id, {
        "day_of_week": cls.day_of_week,
        "session_type": cls.session_type,
        "start_session": cls.start_session,
        "duration": cls.duration,
    })
    if conflict:
        with engine.connect() as conn:
            conflict_info = conn.execute(
                select(
                    classes.c.subject,
                    classes.c.day_of_week,
                    classes.c.session_type,
                    classes.c.start_session,
                    classes.c.duration,
                    teachers.c.full_name.label("teacher_name"),
                    teachers.c.subject_group,
                )
                .join(teachers, classes.c.teacher_id == teachers.c.id)
                .where(classes.c.id == conflict.id)
            ).fetchone()
        subject = (conflict_info.subject or conflict_info.subject_group) if conflict_info else f"Lớp {conflict.id}"
        teacher = conflict_info.teacher_name if conflict_info else ""
        return jsonify(
            ok=False,
            error=f"Lịch của bạn đã bị trùng với môn {subject}.",
            conflict_subject=subject,
            conflict_teacher=teacher,
            conflict_schedule=session_label(
                conflict_info.session_type,
                conflict_info.start_session,
                conflict_info.duration,
            ) + f" – {day_name(conflict_info.day_of_week)}" if conflict_info else "",
        )

    with engine.connect() as conn:
        conn.execute(
            insert(enrollments).values(
                student_id=student_id,
                class_id=class_id,
                enrolled_at=now_vn(),
            )
        )
        conn.commit()

    return jsonify(ok=True)


@app.route("/student/enroll/<int:class_id>", methods=["DELETE"])
@student_required
def student_cancel_enroll(class_id):
    student_id = session["user_id"]
    with engine.connect() as conn:
        conn.execute(
            delete(enrollments).where(
                and_(enrollments.c.student_id == student_id,
                     enrollments.c.class_id == class_id)
            )
        )
        conn.commit()
    return jsonify(ok=True)


@app.route("/api/class-counts")
def api_class_counts():
    is_admin = session.get("is_admin") is True
    user_type = session.get("user_type")
    if not is_admin and not user_type:
        return jsonify({}), 401

    with engine.connect() as conn:
        if is_admin:
            where_clause = classes.c.is_published == 1
        elif user_type == "student":
            student_grade = session.get("grade")
            where_clause = and_(
                classes.c.is_published == 1,
                classes.c.grade == student_grade,
                classes.c.location.isnot(None),
                classes.c.location != '',
                classes.c.max_capacity.isnot(None),
            )
        elif user_type == "teacher":
            teacher_id = session.get("user_id")
            where_clause = classes.c.teacher_id == teacher_id
        else:
            return jsonify({})

        pub_classes = conn.execute(
            select(classes.c.id, classes.c.location, classes.c.max_capacity)
            .where(where_clause)
        ).fetchall()
        counts = {}
        for row in pub_classes:
            cnt = conn.execute(
                select(func.count()).where(enrollments.c.class_id == row.id)
            ).scalar()
            counts[str(row.id)] = {
                "count": cnt,
                "location": row.location or "",
                "max_capacity": row.max_capacity,
            }
    return jsonify(counts)

# ---------------------------------------------------------------------------
# Admin routes
# ---------------------------------------------------------------------------

@app.route("/admin/login", methods=["GET"])
def admin_login_page():
    return render_template("admin/login.html")


@app.route("/admin/login", methods=["POST"])
def admin_login():
    password = request.form.get("password", "")
    admin_pw = get_setting("admin_password", "Admin@123")
    if password == admin_pw:
        session["is_admin"] = True
        return redirect(url_for("admin_index"))
    flash("Sai mật khẩu.")
    return redirect(url_for("admin_login_page"))


@app.route("/admin/logout", methods=["POST"])
def admin_logout():
    session.pop("is_admin", None)
    return redirect(url_for("admin_login_page"))


@app.route("/admin/seed-data", methods=["POST"])
@admin_required
def admin_seed_data():
    try:
        with engine.connect() as conn:
            # Clear all data first so reseed always works
            for tbl in ("enrollments", "classes", "students", "teachers", "rooms"):
                conn.execute(text(f"DELETE FROM {tbl}"))
            # Reset PG sequences to 0 before re-seeding explicit IDs
            if not DATABASE_URL.startswith("sqlite"):
                for tbl in ("teachers", "students", "classes", "enrollments"):
                    conn.execute(text(
                        f"SELECT setval(pg_get_serial_sequence('{tbl}', 'id'), 1, false)"
                    ))
            conn.commit()
            _seed_sample_data(conn)
        return jsonify(ok=True, message="Đã tạo dữ liệu mẫu thành công: 8 giáo viên, 14 học sinh, 19 lớp học, 13 đăng ký, 46 phòng.")
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500


@app.route("/admin/rooms/upload", methods=["POST"])
@admin_required
def admin_rooms_upload():
    f = request.files.get("file")
    if not f:
        return jsonify(ok=False, error="Không có file.")
    try:
        import openpyxl as _openpyxl
        wb = _openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        name_col = None
        for i, h in enumerate(headers):
            if "PHÒNG" in h.upper() or "TÊN" in h.upper():
                name_col = i
                break
        if name_col is None:
            return jsonify(ok=False, error="Không tìm thấy cột tên phòng.")
        added = skipped = 0
        with engine.connect() as conn:
            for row in ws.iter_rows(min_row=2, values_only=True):
                name = str(row[name_col] or "").strip()
                if not name:
                    continue
                try:
                    conn.execute(insert(rooms).values(name=name))
                    conn.commit()
                    added += 1
                except Exception:
                    conn.rollback()
                    skipped += 1
        return jsonify(ok=True, added=added, skipped=skipped)
    except Exception as e:
        return jsonify(ok=False, error=str(e))


@app.route("/admin/rooms/clear", methods=["POST"])
@admin_required
def admin_rooms_clear():
    with engine.connect() as conn:
        conn.execute(text("DELETE FROM rooms"))
        conn.commit()
    return jsonify(ok=True)


@app.route("/admin/rooms/delete", methods=["POST"])
@admin_required
def admin_rooms_delete_one():
    room_name = (request.get_json(force=True).get("room_name") or "").strip()
    if not room_name:
        return jsonify(ok=False, error="Thiếu tên phòng.")
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM rooms WHERE name = :n"), {"n": room_name})
        conn.execute(
            delete(room_external_busy).where(room_external_busy.c.room_name == room_name)
        )
    _bump(event_type="schedule")
    return jsonify(ok=True)


# Header columns in the busy-rooms Excel (cols 4-15, 1-indexed):
# Sáng T2, Sáng T3, Sáng T4, Sáng T5, Sáng T6, Sáng T7,
# Chiều T2, Chiều T3, Chiều T4, Chiều T5, Chiều T6, Chiều T7
_BUSY_COL_MAP = [
    ("morning", 2), ("morning", 3), ("morning", 4),
    ("morning", 5), ("morning", 6), ("morning", 7),
    ("afternoon", 2), ("afternoon", 3), ("afternoon", 4),
    ("afternoon", 5), ("afternoon", 6), ("afternoon", 7),
]


@app.route("/admin/rooms/busy-upload", methods=["POST"])
@admin_required
def admin_rooms_busy_upload():
    f = request.files.get("file")
    if not f:
        return jsonify(ok=False, error="Không có file"), 400
    try:
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active

        # Detect column mapping from header row
        # Find which columns correspond to the 12 sessions
        header = [str(c).strip() if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        # Map column index → (session_type, day_of_week)
        # Try to detect from header, fallback to fixed positions
        col_map = {}  # 0-indexed col → (session_type, day)
        for ci, h in enumerate(header):
            h_low = h.lower()
            ses = "morning" if "sáng" in h_low or "sang" in h_low else \
                  "afternoon" if "chiều" in h_low or "chieu" in h_low else None
            for dow, kw in [(2, "2"), (3, "3"), (4, "4"), (5, "5"), (6, "6"), (7, "7")]:
                if kw in h:
                    if ses:
                        col_map[ci] = (ses, dow)
                    break

        # Fallback: if header detection fails, use fixed col positions (cols 3-14, 0-indexed)
        if len(col_map) < 12:
            col_map = {3 + i: v for i, v in enumerate(_BUSY_COL_MAP)}

        added = 0
        current_room = None
        rows_data = list(ws.iter_rows(min_row=2, values_only=True))
        seen_rooms = []

        def _parse_grades(raw):
            """'10 - 11' / '10-11' → '10,11'; '10' → '10'; 'x'/'X' → None (busy); '' → skip"""
            s = str(raw).strip().lower()
            if not s or s == "none":
                return "skip"          # blank → phòng trống, không giới hạn khối
            if s in ("x",):
                return None            # busy
            # Normalise separators: '10 - 11', '10-11', '10,11'
            import re
            nums = re.findall(r'\d+', s)
            if nums:
                return ",".join(nums)  # '10,11' or '10' or '11'
            return None                # unrecognised → treat as busy

        with engine.begin() as conn:
            conn.execute(text("DELETE FROM room_external_busy"))
            conn.execute(text("DELETE FROM room_grade_slots"))
            for row in rows_data:
                # Column B (index 1) = room name
                if row[1] and str(row[1]).strip():
                    current_room = str(row[1]).strip()
                    if current_room not in seen_rooms:
                        seen_rooms.append(current_room)
                # Column C (index 2) = tiết
                tiet_val = row[2]
                if not current_room or not tiet_val:
                    continue
                try:
                    tiet = int(tiet_val)
                except (ValueError, TypeError):
                    continue
                if tiet not in (1, 2, 3, 4):
                    continue

                for ci, (ses, dow) in col_map.items():
                    raw = row[ci] if ci < len(row) else None
                    if raw is None:
                        continue
                    grades = _parse_grades(raw)
                    if grades == "skip":
                        continue           # phòng trống không hạn chế → bỏ qua
                    elif grades is None:
                        conn.execute(insert(room_external_busy).values(
                            room_name=current_room, day_of_week=dow,
                            session_type=ses, tiet=tiet,
                        ))
                        added += 1
                    else:
                        conn.execute(insert(room_grade_slots).values(
                            room_name=current_room, day_of_week=dow,
                            session_type=ses, tiet=tiet,
                            available_grades=grades,
                        ))
                        added += 1

            # Sync room names into rooms table
            for rn in seen_rooms:
                conn.execute(text("INSERT INTO rooms (name) VALUES (:n) ON CONFLICT (name) DO NOTHING"), {"n": rn})

        _bump(event_type="schedule")
        return jsonify(ok=True, added=added, rooms_synced=len(seen_rooms))
    except Exception as e:
        return jsonify(ok=False, error=str(e))


@app.route("/admin/rooms/busy-clear", methods=["POST"])
@admin_required
def admin_rooms_busy_clear():
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM room_external_busy"))
        conn.execute(text("DELETE FROM room_grade_slots"))
    return jsonify(ok=True)


@app.route("/admin/rooms/schedule")
@admin_required
def admin_rooms_schedule():
    with engine.connect() as conn:
        all_rooms = [r.name for r in conn.execute(
            select(rooms).order_by(rooms.c.name)
        ).fetchall()]
        busy_rows = conn.execute(select(room_external_busy)).fetchall()
        grade_rows = conn.execute(select(room_grade_slots)).fetchall()
    busy_map = {}
    for r in busy_rows:
        key = f"{r.day_of_week}_{r.session_type}_{r.tiet}"
        busy_map.setdefault(r.room_name, []).append(key)
    grade_map = {}
    for r in grade_rows:
        key = f"{r.day_of_week}_{r.session_type}_{r.tiet}"
        if r.room_name not in grade_map:
            grade_map[r.room_name] = {}
        grade_map[r.room_name][key] = r.available_grades
    return jsonify(rooms=all_rooms, busy=busy_map, grade_map=grade_map)


@app.route("/admin/rooms/schedule", methods=["POST"])
@admin_required
def admin_rooms_schedule_save():
    data = request.get_json(force=True)
    room_list  = [str(r).strip() for r in data.get("rooms", []) if str(r).strip()]
    busy_map   = data.get("busy", {})
    grade_map  = data.get("grade_map", {})  # {room: {key: "10,11"}}
    VALID_SESS = ("morning", "afternoon")
    VALID_TIET = (1, 2, 3, 4)
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM rooms"))
        conn.execute(text("DELETE FROM room_external_busy"))
        conn.execute(text("DELETE FROM room_grade_slots"))
        for rn in room_list:
            conn.execute(text("INSERT INTO rooms (name) VALUES (:n)"), {"n": rn})
            for key in busy_map.get(rn, []):
                parts = key.split("_", 2)
                if len(parts) != 3:
                    continue
                try:
                    dow, ses, tiet = int(parts[0]), parts[1], int(parts[2])
                except ValueError:
                    continue
                if dow not in range(2, 8) or ses not in VALID_SESS or tiet not in VALID_TIET:
                    continue
                conn.execute(insert(room_external_busy).values(
                    room_name=rn, day_of_week=dow, session_type=ses, tiet=tiet,
                ))
            for key, grades in (grade_map.get(rn) or {}).items():
                parts = key.split("_", 2)
                if len(parts) != 3:
                    continue
                try:
                    dow, ses, tiet = int(parts[0]), parts[1], int(parts[2])
                except ValueError:
                    continue
                if dow not in range(2, 8) or ses not in VALID_SESS or tiet not in VALID_TIET:
                    continue
                conn.execute(insert(room_grade_slots).values(
                    room_name=rn, day_of_week=dow, session_type=ses,
                    tiet=tiet, available_grades=str(grades),
                ))
    _bump(event_type="schedule")
    return jsonify(ok=True)


@app.route("/admin/rooms/add-manual", methods=["POST"])
@admin_required
def admin_rooms_add_manual():
    data = request.get_json(force=True)
    room_name = (data.get("room_name") or "").strip()
    if not room_name:
        return jsonify(ok=False, error="Tên phòng không được để trống.")
    busy = data.get("busy", [])  # [{day_of_week, session_type, tiet}, ...]

    with engine.begin() as conn:
        conn.execute(text("INSERT INTO rooms (name) VALUES (:n) ON CONFLICT (name) DO NOTHING"), {"n": room_name})
        conn.execute(delete(room_external_busy).where(room_external_busy.c.room_name == room_name))
        for slot in busy:
            try:
                dow  = int(slot["day_of_week"])
                ses  = slot["session_type"]
                tiet = int(slot["tiet"])
            except (KeyError, ValueError, TypeError):
                continue
            if dow not in range(2, 8) or ses not in ("morning", "afternoon") or tiet not in (1, 2, 3, 4):
                continue
            conn.execute(insert(room_external_busy).values(
                room_name=room_name,
                day_of_week=dow,
                session_type=ses,
                tiet=tiet,
            ))
    return jsonify(ok=True)


@app.route("/admin/rooms/busy-template")
@admin_required
def admin_rooms_busy_template():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io as _io

    wb = Workbook()
    ws = wb.active
    ws.title = "Lịch phòng bận"

    header_fill  = PatternFill("solid", fgColor="1E40AF")
    morning_fill = PatternFill("solid", fgColor="DBEAFE")
    afternoon_fill = PatternFill("solid", fgColor="FEF3C7")
    bold_white = Font(bold=True, color="FFFFFF")
    bold_dark  = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="9CA3AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: merged headers
    ws.merge_cells("A1:A2"); ws["A1"] = "STT"
    ws.merge_cells("B1:B2"); ws["B1"] = "Phòng"
    ws.merge_cells("C1:C2"); ws["C1"] = "Tiết"
    ws.merge_cells("D1:I1"); ws["D1"] = "Buổi Sáng"
    ws.merge_cells("J1:O1"); ws["J1"] = "Buổi Chiều"

    # Row 2: day headers
    days = ["T2", "T3", "T4", "T5", "T6", "T7"]
    for i, d in enumerate(days):
        ws.cell(row=2, column=4 + i).value  = d  # morning
        ws.cell(row=2, column=10 + i).value = d  # afternoon

    # Style row 1 + 2 headers
    for col in range(1, 16):
        c1 = ws.cell(row=1, column=col)
        c2 = ws.cell(row=2, column=col)
        c1.fill = header_fill; c1.font = bold_white; c1.alignment = center; c1.border = border
        c2.fill = header_fill; c2.font = bold_white; c2.alignment = center; c2.border = border

    # Sample data rows (2 rooms × 4 tiết)
    sample_rooms = ["A101", "A102"]
    row = 3
    for stt, room in enumerate(sample_rooms, start=1):
        first_row = row
        for tiet in (1, 2, 3, 4):
            ws.cell(row=row, column=3).value = tiet
            # Style morning cols
            for col in range(4, 10):
                c = ws.cell(row=row, column=col)
                c.fill = morning_fill; c.alignment = center; c.border = border
            # Style afternoon cols
            for col in range(10, 16):
                c = ws.cell(row=row, column=col)
                c.fill = afternoon_fill; c.alignment = center; c.border = border
            # Example: mark sample busy cell
            if tiet == 1 and stt == 1:
                ws.cell(row=row, column=4).value = "x"  # Sáng T2
            row += 1
        # Merge STT and Phòng across 4 rows for this room
        last_row = row - 1
        ws.merge_cells(f"A{first_row}:A{last_row}")
        ws.merge_cells(f"B{first_row}:B{last_row}")
        ws[f"A{first_row}"] = stt
        ws[f"B{first_row}"] = room
        for col in (1, 2):
            c = ws.cell(row=first_row, column=col)
            c.alignment = center; c.border = border

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 6
    for col in range(4, 16):
        ws.column_dimensions[get_column_letter(col)].width = 7

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="mau_phong_ban.xlsx",
    )


@app.route("/api/slot-impact-grid")
@teacher_required
def api_slot_impact_grid():
    """Return combo counts for ALL valid slots at once (for mini slot-picker grid)."""
    grade = request.args.get("grade", type=int)
    dur   = request.args.get("duration", type=int)
    if grade not in (10, 11, 12):
        return jsonify(error="Khối không hợp lệ"), 400
    if dur not in (2, 4):
        return jsonify(error="Số tiết phải là 2 hoặc 4"), 400

    teacher_id = session["user_id"]
    subject = _norm_subj(session.get("subject_group") or "")
    by_subj = _build_by_subj(grade)
    max_combos = count_combos_max(grade, subject)

    # Build set of slots already occupied by this teacher
    with engine.connect() as conn:
        teacher_classes = conn.execute(
            select(classes.c.day_of_week, classes.c.session_type,
                   classes.c.start_session, classes.c.duration)
            .where(classes.c.teacher_id == teacher_id)
        ).fetchall()
    teacher_occupied = set()
    for c in teacher_classes:
        end = c.start_session + c.duration - 1
        for t in range(c.start_session, end + 1):
            teacher_occupied.add((c.day_of_week, c.session_type, t))

    def _teacher_conflict(dow, ses, start):
        """True nếu GV đã có lớp chồng tiết với slot [start, start+dur-1]."""
        end = start + dur - 1
        return any((dow, ses, t) in teacher_occupied for t in range(start, end + 1))

    def _combos_for_slot(dow, ses, start):
        new_slot = {"day_of_week": dow, "session_type": ses,
                    "start_session": start, "duration": dur}
        tmp = {k: list(v) for k, v in by_subj.items()}
        tmp.setdefault(subject, []).append(new_slot)
        others = sorted(s for s in tmp if s != subject)
        count = [0]
        _backtrack(tmp, others, 0, [new_slot], count, 100)
        return count[0]

    valid_starts = [1, 3] if dur == 2 else [1]
    grid = {}
    for dow in range(2, 8):
        for ses in ("morning", "afternoon"):
            for start in range(1, 5):
                key = f"{dow}_{ses}_{start}"
                if start not in valid_starts:
                    grid[key] = -1  # not an allowed starting tiet
                elif _teacher_conflict(dow, ses, start):
                    grid[key] = -1  # GV đã có lớp ở khung giờ này
                else:
                    grid[key] = _combos_for_slot(dow, ses, start)
    return jsonify(grid=grid, max_combos=max_combos)


@app.route("/api/available-rooms")
def api_available_rooms():
    if not (session.get("is_admin") or session.get("user_type") == "teacher"):
        return jsonify([]), 401
    try:
        dow   = int(request.args["day_of_week"])
        st    = request.args["session_type"]
        ss    = int(request.args["start_session"])
        dur   = int(request.args["duration"])
        grade = int(request.args.get("grade", 0))
    except (KeyError, ValueError):
        return jsonify([]), 400
    end = ss + dur - 1
    with engine.connect() as conn:
        all_rooms = sorted([r.name for r in conn.execute(select(rooms)).fetchall()], key=_room_sort_key)
        # rooms already booked for overlapping slots
        booked = conn.execute(
            select(classes.c.location).where(
                and_(
                    classes.c.day_of_week == dow,
                    classes.c.session_type == st,
                    classes.c.location.isnot(None),
                    classes.c.start_session <= end,
                    (classes.c.start_session + classes.c.duration - 1) >= ss,
                )
            )
        ).fetchall()
        booked_set = {r.location for r in booked if r.location}
        # rooms marked busy externally (any tiết in the requested range)
        ext_busy = conn.execute(
            select(room_external_busy.c.room_name).where(
                and_(
                    room_external_busy.c.day_of_week == dow,
                    room_external_busy.c.session_type == st,
                    room_external_busy.c.tiet >= ss,
                    room_external_busy.c.tiet <= end,
                )
            )
        ).fetchall()
        ext_busy_set = {r.room_name for r in ext_busy}
        # rooms grade-restricted and not allowed for this grade
        grade_blocked_set = set()
        if grade:
            grade_rows = conn.execute(
                select(room_grade_slots.c.room_name, room_grade_slots.c.available_grades).where(
                    and_(
                        room_grade_slots.c.day_of_week == dow,
                        room_grade_slots.c.session_type == st,
                        room_grade_slots.c.tiet >= ss,
                        room_grade_slots.c.tiet <= end,
                    )
                )
            ).fetchall()
            for r in grade_rows:
                allowed = [g.strip() for g in r.available_grades.split(",")]
                if str(grade) not in allowed:
                    grade_blocked_set.add(r.room_name)
    available = [r for r in all_rooms
                 if r not in booked_set and r not in ext_busy_set and r not in grade_blocked_set]
    return jsonify(available)


@app.route("/api/slot-impact")
@teacher_required
def api_slot_impact():
    try:
        grade = int(request.args["grade"])
        dow   = int(request.args["day_of_week"])
        sess  = request.args["session_type"]
        start = int(request.args["start_session"])
        dur   = int(request.args["duration"])
    except (KeyError, ValueError, TypeError):
        return jsonify(error="Params không hợp lệ"), 400
    if grade not in (10, 11, 12):
        return jsonify(error="Khối không hợp lệ"), 400
    if sess not in ("morning", "afternoon"):
        return jsonify(error="Buổi không hợp lệ"), 400
    if not (1 <= start <= 4) or not (1 <= dur <= 4) or start + dur - 1 > 4:
        return jsonify(error="Tiết không hợp lệ"), 400

    subject = session.get("subject_group")
    new_cls = {"subject": subject, "day_of_week": dow, "session_type": sess,
               "start_session": start, "duration": dur}
    combos = count_combos_including(grade, new_cls)
    return jsonify(combos=combos, cap=100)


@app.route("/api/room-schedule")
def api_room_schedule():
    if not (session.get("is_admin") or session.get("user_type") == "teacher"):
        return jsonify({}), 401
    with engine.connect() as conn:
        all_rooms = sorted(
            [r.name for r in conn.execute(select(rooms)).fetchall()],
            key=_room_sort_key,
        )
        rows = conn.execute(
            select(
                classes.c.location,
                classes.c.day_of_week,
                classes.c.session_type,
                classes.c.start_session,
                classes.c.duration,
                classes.c.grade,
                classes.c.subject,
                teachers.c.full_name.label("teacher_name"),
            )
            .join(teachers, classes.c.teacher_id == teachers.c.id)
            .where(classes.c.location.isnot(None))
        ).fetchall()
        ext_busy_rows = conn.execute(select(room_external_busy)).fetchall()
        grade_slot_rows = conn.execute(select(room_grade_slots)).fetchall()
    bookings = [
        {
            "room": r.location,
            "day": r.day_of_week,
            "session": r.session_type,
            "start": r.start_session,
            "duration": r.duration,
            "label": (
                f"Khối {r.grade}"
                + (f" - {r.subject}" if r.subject else "")
                + f" ({r.teacher_name})"
            ),
            "type": "class",
        }
        for r in rows
    ]
    for r in ext_busy_rows:
        bookings.append({
            "room": r.room_name,
            "day": r.day_of_week,
            "session": r.session_type,
            "start": r.tiet,
            "duration": 1,
            "label": "Bận (lịch ngoài)",
            "type": "external",
        })
    grade_slots = [
        {
            "room": r.room_name,
            "day": r.day_of_week,
            "session": r.session_type,
            "tiet": r.tiet,
            "grades": r.available_grades,  # "10", "11", "10,11"
        }
        for r in grade_slot_rows
    ]
    return jsonify({"rooms": all_rooms, "bookings": bookings, "grade_slots": grade_slots})


@app.route("/admin")
@admin_required
def admin_index():
    with engine.connect() as conn:
        teacher_count = conn.execute(select(func.count()).select_from(teachers)).scalar()
        student_count = conn.execute(select(func.count()).select_from(students)).scalar()
        class_count = conn.execute(select(func.count()).select_from(classes)).scalar()
        enrollment_count = conn.execute(select(func.count()).select_from(enrollments)).scalar()
        teacher_active = conn.execute(
            select(func.count()).select_from(teachers).where(
                and_(teachers.c.is_first_login == 0, teachers.c.password_hash.isnot(None))
            )
        ).scalar()
        student_active = conn.execute(
            select(func.count()).select_from(students).where(
                and_(students.c.is_first_login == 0, students.c.password_hash.isnot(None))
            )
        ).scalar()

        teacher_list = conn.execute(
            select(teachers).order_by(teachers.c.full_name)
        ).fetchall()
        student_list = conn.execute(
            select(students).order_by(students.c.grade, students.c.class_name, students.c.full_name)
        ).fetchall()
        room_list  = sorted(conn.execute(select(rooms)).fetchall(), key=lambda r: _room_sort_key(r.name))
        room_count = len(room_list)

    teacher_reg_open = get_setting("teacher_reg_open", "0") == "1"
    student_reg_open = get_setting("student_reg_open", "0") == "1"
    maintenance = get_setting("maintenance_mode", "0") == "1"
    schedule_constraint = get_setting("schedule_constraint", "1") == "1"
    with engine.connect() as conn2:
        busy_room_count = conn2.execute(
            select(func.count()).select_from(room_external_busy)
        ).scalar() or 0
    return render_template(
        "admin/index.html",
        stats={
            "teachers": teacher_count,
            "students": student_count,
            "classes": class_count,
            "enrollments": enrollment_count,
            "teacher_active": teacher_active,
            "student_active": student_active,
        },
        teacher_list=teacher_list,
        student_list=student_list,
        teacher_reg_open=teacher_reg_open,
        student_reg_open=student_reg_open,
        maintenance=maintenance,
        schedule_constraint=schedule_constraint,
        room_list=room_list,
        room_count=room_count,
        busy_room_count=busy_room_count,
    )

# --- Admin: Teacher management ---

@app.route("/admin/teachers/upload", methods=["POST"])
@admin_required
def admin_teachers_upload():
    file = request.files.get("file")
    if not file:
        flash("Chưa chọn file.")
        return redirect(url_for("admin_index"))

    wb = openpyxl.load_workbook(file)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    required = ["Họ và tên", "Mã đăng nhập", "Giới tính", "Tổ bộ môn"]
    for r in required:
        if r not in headers:
            flash(f"Thiếu cột: {r}")
            return redirect(url_for("admin_index"))

    idx = {h: headers.index(h) for h in required}
    count = 0
    with engine.connect() as conn:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            full_name = str(row[idx["Họ và tên"]] or "").strip()
            cccd = str(row[idx["Mã đăng nhập"]] or "").strip()
            gender = str(row[idx["Giới tính"]] or "").strip()
            subject_group = str(row[idx["Tổ bộ môn"]] or "").strip()
            if not full_name or not cccd:
                continue

            existing = conn.execute(
                select(teachers).where(teachers.c.cccd == cccd)
            ).fetchone()
            if existing:
                conn.execute(
                    update(teachers).where(teachers.c.cccd == cccd).values(
                        full_name=full_name,
                        gender=gender,
                        subject_group=subject_group,
                    )
                )
            else:
                conn.execute(
                    insert(teachers).values(
                        full_name=full_name,
                        cccd=cccd,
                        gender=gender,
                        subject_group=subject_group,
                        email=None,
                        password_hash=None,
                        is_first_login=1,
                    )
                )
            count += 1
        conn.commit()

    flash(f"Đã nhập {count} giáo viên.")
    return redirect(url_for("admin_index"))


@app.route("/admin/teachers/template")
@admin_required
def admin_teachers_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Giáo viên"
    ws.append(["Họ và tên", "Mã đăng nhập", "Giới tính", "Tổ bộ môn"])
    ws.append(["Nguyễn Văn A", "012345678901", "Nam", "Toán"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="mau_giao_vien.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/teachers/add", methods=["POST"])
@admin_required
def admin_teachers_add():
    full_name = request.form.get("full_name", "").strip()
    cccd = request.form.get("cccd", "").strip()
    gender = request.form.get("gender", "").strip()
    subject_group = request.form.get("subject_group", "").strip()

    if not full_name or not cccd:
        flash("Vui lòng nhập đầy đủ thông tin.")
        return redirect(url_for("admin_index"))

    with engine.connect() as conn:
        existing = conn.execute(
            select(teachers).where(teachers.c.cccd == cccd)
        ).fetchone()
        if existing:
            conn.execute(
                update(teachers).where(teachers.c.cccd == cccd).values(
                    full_name=full_name,
                    gender=gender,
                    subject_group=subject_group,
                )
            )
            flash("Đã cập nhật giáo viên.")
        else:
            conn.execute(
                insert(teachers).values(
                    full_name=full_name,
                    cccd=cccd,
                    gender=gender,
                    subject_group=subject_group,
                    email=None,
                    password_hash=None,
                    is_first_login=1,
                )
            )
            flash("Đã thêm giáo viên.")
        conn.commit()

    return redirect(url_for("admin_index"))


@app.route("/admin/teachers/<int:teacher_id>/reset", methods=["POST"])
@admin_required
def admin_teachers_reset(teacher_id):
    with engine.connect() as conn:
        # Get all classes for this teacher
        teacher_classes = conn.execute(
            select(classes.c.id).where(classes.c.teacher_id == teacher_id)
        ).fetchall()
        class_ids = [c.id for c in teacher_classes]

        # Delete enrollments for those classes
        if class_ids:
            conn.execute(
                delete(enrollments).where(enrollments.c.class_id.in_(class_ids))
            )
            conn.execute(
                delete(classes).where(classes.c.teacher_id == teacher_id)
            )

        # Reset teacher account
        conn.execute(
            update(teachers).where(teachers.c.id == teacher_id).values(
                email=None,
                password_hash=None,
                is_first_login=1,
            )
        )
        conn.commit()

    flash("Đã reset tài khoản giáo viên.")
    return redirect(url_for("admin_index"))


@app.route("/admin/teachers/<int:teacher_id>/reset-password", methods=["POST"])
@admin_required
def admin_teachers_reset_password(teacher_id):
    with engine.connect() as conn:
        teacher = conn.execute(
            select(teachers.c.id, teachers.c.full_name, teachers.c.is_first_login)
            .where(teachers.c.id == teacher_id)
        ).fetchone()
    if not teacher:
        return jsonify(ok=False, error="Không tìm thấy giáo viên.")
    if teacher.is_first_login:
        return jsonify(ok=False, error="Tài khoản chưa kích hoạt, không cần khôi phục mật khẩu.")

    temp_pw = generate_temp_password()
    with engine.begin() as conn:
        conn.execute(
            update(teachers).where(teachers.c.id == teacher_id).values(
                password_hash=generate_password_hash(temp_pw),
                must_change_password=1,
            )
        )
    return jsonify(ok=True, new_password=temp_pw, name=teacher.full_name)


@app.route("/admin/teachers/<int:teacher_id>", methods=["DELETE"])
@admin_required
def admin_teachers_delete(teacher_id):
    with engine.connect() as conn:
        teacher_classes = conn.execute(
            select(classes.c.id).where(classes.c.teacher_id == teacher_id)
        ).fetchall()
        class_ids = [c.id for c in teacher_classes]

        if class_ids:
            conn.execute(
                delete(enrollments).where(enrollments.c.class_id.in_(class_ids))
            )
            conn.execute(
                delete(classes).where(classes.c.teacher_id == teacher_id)
            )

        conn.execute(delete(teachers).where(teachers.c.id == teacher_id))
        conn.commit()

    return jsonify(ok=True)

@app.route("/admin/teachers/clear", methods=["POST"])
@admin_required
def admin_teachers_clear():
    with engine.begin() as conn:
        all_class_ids = [r.id for r in conn.execute(select(classes.c.id)).fetchall()]
        if all_class_ids:
            conn.execute(delete(enrollments).where(enrollments.c.class_id.in_(all_class_ids)))
        conn.execute(delete(classes))
        conn.execute(delete(teachers))
    flash("Đã xóa toàn bộ danh sách giáo viên.", "success")
    return redirect(url_for("admin_index"))


# --- Admin: Student management ---

@app.route("/admin/students/upload", methods=["POST"])
@admin_required
def admin_students_upload():
    file = request.files.get("file")
    if not file:
        flash("Chưa chọn file.")
        return redirect(url_for("admin_index"))

    wb = openpyxl.load_workbook(file)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    required = ["Họ và tên", "Mã đăng nhập", "Lớp", "Khối"]
    for r in required:
        if r not in headers:
            flash(f"Thiếu cột: {r}")
            return redirect(url_for("admin_index"))

    idx = {h: headers.index(h) for h in required}
    count = 0
    with engine.connect() as conn:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            full_name = str(row[idx["Họ và tên"]] or "").strip()
            cccd = str(row[idx["Mã đăng nhập"]] or "").strip()
            class_name = str(row[idx["Lớp"]] or "").strip()
            try:
                grade = int(row[idx["Khối"]] or 0)
            except (ValueError, TypeError):
                grade = 0
            if not full_name or not cccd:
                continue

            existing = conn.execute(
                select(students).where(students.c.cccd == cccd)
            ).fetchone()
            if existing:
                conn.execute(
                    update(students).where(students.c.cccd == cccd).values(
                        full_name=full_name,
                        class_name=class_name,
                        grade=grade,
                    )
                )
            else:
                conn.execute(
                    insert(students).values(
                        full_name=full_name,
                        cccd=cccd,
                        class_name=class_name,
                        grade=grade,
                    )
                )
            count += 1
        conn.commit()

    flash(f"Đã nhập {count} học sinh.")
    return redirect(url_for("admin_index"))


@app.route("/admin/students/template")
@admin_required
def admin_students_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Học sinh"
    ws.append(["Họ và tên", "Mã đăng nhập", "Lớp", "Khối"])
    ws.append(["Trần Thị B", "098765432109", "10A1", 10])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="mau_hoc_sinh.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/students/add", methods=["POST"])
@admin_required
def admin_students_add():
    full_name = request.form.get("full_name", "").strip()
    cccd = request.form.get("cccd", "").strip()
    class_name = request.form.get("class_name", "").strip()
    try:
        grade = int(request.form.get("grade", 0))
    except ValueError:
        grade = 0

    if not full_name or not cccd:
        flash("Vui lòng nhập đầy đủ thông tin.")
        return redirect(url_for("admin_index"))

    with engine.connect() as conn:
        existing = conn.execute(
            select(students).where(students.c.cccd == cccd)
        ).fetchone()
        if existing:
            conn.execute(
                update(students).where(students.c.cccd == cccd).values(
                    full_name=full_name,
                    class_name=class_name,
                    grade=grade,
                )
            )
            flash("Đã cập nhật học sinh.")
        else:
            conn.execute(
                insert(students).values(
                    full_name=full_name,
                    cccd=cccd,
                    class_name=class_name,
                    grade=grade,
                )
            )
            flash("Đã thêm học sinh.")
        conn.commit()

    return redirect(url_for("admin_index"))


@app.route("/admin/students/<int:student_id>/reset", methods=["POST"])
@admin_required
def admin_students_reset(student_id):
    with engine.begin() as conn:
        conn.execute(delete(enrollments).where(enrollments.c.student_id == student_id))
        conn.execute(
            update(students).where(students.c.id == student_id).values(
                password_hash=None, is_first_login=1
            )
        )
    flash("Đã reset tài khoản và xóa đăng ký của học sinh.", "success")
    return redirect(url_for("admin_index"))


@app.route("/admin/students/<int:student_id>/reset-password", methods=["POST"])
@admin_required
def admin_students_reset_password(student_id):
    with engine.connect() as conn:
        student = conn.execute(
            select(students.c.id, students.c.full_name, students.c.is_first_login)
            .where(students.c.id == student_id)
        ).fetchone()
    if not student:
        return jsonify(ok=False, error="Không tìm thấy học sinh.")
    if student.is_first_login:
        return jsonify(ok=False, error="Tài khoản chưa kích hoạt, không cần khôi phục mật khẩu.")

    temp_pw = generate_temp_password()
    with engine.begin() as conn:
        conn.execute(
            update(students).where(students.c.id == student_id).values(
                password_hash=generate_password_hash(temp_pw),
                must_change_password=1,
            )
        )
    return jsonify(ok=True, new_password=temp_pw, name=student.full_name)


@app.route("/admin/students/<int:student_id>", methods=["DELETE"])
@admin_required
def admin_students_delete(student_id):
    with engine.connect() as conn:
        conn.execute(
            delete(enrollments).where(enrollments.c.student_id == student_id)
        )
        conn.execute(delete(students).where(students.c.id == student_id))
        conn.commit()
    return jsonify(ok=True)


@app.route("/admin/students/clear", methods=["POST"])
@admin_required
def admin_students_clear():
    with engine.begin() as conn:
        conn.execute(delete(enrollments))
        conn.execute(delete(students))
    flash("Đã xóa toàn bộ danh sách học sinh.", "success")
    return redirect(url_for("admin_index"))


# ---------------------------------------------------------------------------
# Operator account management
# ---------------------------------------------------------------------------

@app.route("/admin/operators")
@admin_required
def admin_operators():
    with engine.connect() as conn:
        ops = conn.execute(select(operators).order_by(operators.c.id)).fetchall()
    return render_template("admin/operators.html", operators=ops)


@app.route("/admin/operators", methods=["POST"])
@admin_required
def admin_operators_create():
    full_name  = (request.form.get("full_name") or "").strip()
    login_code = (request.form.get("login_code") or "").strip()
    password   = (request.form.get("password") or "").strip()
    if not full_name or not login_code or not password:
        flash("Vui lòng nhập đầy đủ Họ tên, Mã đăng nhập và Mật khẩu.", "danger")
        return redirect(url_for("admin_operators"))
    pw_hash = generate_password_hash(password)
    try:
        with engine.begin() as conn:
            conn.execute(insert(operators).values(
                full_name=full_name,
                login_code=login_code,
                password_hash=pw_hash,
            ))
        flash(f"Đã tạo tài khoản '{full_name}'.", "success")
    except Exception:
        flash("Mã đăng nhập đã tồn tại.", "danger")
    return redirect(url_for("admin_operators"))


@app.route("/admin/operators/<int:op_id>/reset-password", methods=["POST"])
@admin_required
def admin_operators_reset_pw(op_id):
    password = (request.form.get("password") or "").strip()
    if not password:
        return jsonify(ok=False, error="Mật khẩu không được để trống.")
    with engine.begin() as conn:
        conn.execute(update(operators).where(operators.c.id == op_id).values(
            password_hash=generate_password_hash(password)
        ))
    return jsonify(ok=True)


@app.route("/admin/operators/<int:op_id>", methods=["DELETE"])
@admin_required
def admin_operators_delete(op_id):
    with engine.begin() as conn:
        conn.execute(delete(operators).where(operators.c.id == op_id))
    return jsonify(ok=True)


# --- Admin: Class registration management ---

@app.route("/admin/room-grid")
@admin_required
def admin_room_grid():
    with engine.connect() as conn:
        all_rooms = sorted([r.name for r in conn.execute(select(rooms)).fetchall()], key=_room_sort_key)
        rows = conn.execute(
            select(classes, teachers.c.full_name.label("teacher_name"), teachers.c.subject_group)
            .join(teachers, classes.c.teacher_id == teachers.c.id)
            .where(classes.c.location.isnot(None))
        ).fetchall()
        ext_busy_rows = conn.execute(select(room_external_busy)).fetchall()
        grade_slot_rows = conn.execute(select(room_grade_slots)).fetchall()
    return jsonify(
        rooms=all_rooms,
        bookings=[{
            "class_id": r.id,
            "room": r.location,
            "day": r.day_of_week,
            "session": r.session_type,
            "start": r.start_session,
            "duration": r.duration,
            "grade": r.grade,
            "subject": r.subject or r.subject_group or "",
            "teacher": r.teacher_name,
        } for r in rows],
        ext_busy=[{
            "room": r.room_name, "day": r.day_of_week,
            "session": r.session_type, "tiet": r.tiet,
        } for r in ext_busy_rows],
        grade_slots=[{
            "room": r.room_name, "day": r.day_of_week,
            "session": r.session_type, "tiet": r.tiet, "grades": r.available_grades,
        } for r in grade_slot_rows],
    )


def _teachers_json_list(conn):
    all_teachers = conn.execute(select(teachers).order_by(teachers.c.full_name)).fetchall()
    return [{"id": t.id, "full_name": t.full_name, "subject_group": t.subject_group or ""} for t in all_teachers]


@app.route("/admin/room-detail")
@admin_required
def admin_room_detail():
    with engine.connect() as conn:
        tj = _teachers_json_list(conn)
    return render_template("admin/room_grid.html",
                           teachers_json=tj,
                           is_operator_session=False,
                           api_room_grid=url_for("admin_room_grid"),
                           api_register_class=url_for("admin_register_class"),
                           api_class_base="/admin/classes",
                           api_export_url=url_for("admin_room_detail_export"))


@app.route("/admin/room-detail/export")
@admin_required
def admin_room_detail_export():
    with engine.connect() as conn:
        all_rooms = sorted([r.name for r in conn.execute(select(rooms)).fetchall()], key=_room_sort_key)
        rows = conn.execute(
            select(classes, teachers.c.full_name.label("teacher_name"), teachers.c.subject_group)
            .join(teachers, classes.c.teacher_id == teachers.c.id)
            .where(classes.c.location.isnot(None))
        ).fetchall()
        ext_busy_rows = conn.execute(select(room_external_busy)).fetchall()

    DAYS   = [2, 3, 4, 5, 6, 7]
    DNAMES = {2:"Thứ Hai", 3:"Thứ Ba", 4:"Thứ Tư", 5:"Thứ Năm", 6:"Thứ Sáu", 7:"Thứ Bảy"}
    SLOTS  = [1, 3]
    SLBL   = {1:"T1-2", 3:"T3-4"}

    # Build booking lookup: (session, room, day, slot_start) → booking row
    book_map = {}
    for r in rows:
        gs = 1 if r.start_session <= 2 else 3
        k = (r.session_type, r.location, r.day_of_week, gs)
        if k not in book_map:
            book_map[k] = r

    busy_set = set()
    for e in ext_busy_rows:
        busy_set.add((e.session_type, e.room_name, e.day_of_week, e.tiet))

    def is_ext_busy(ses, room, day, slot_start):
        return any((ses, room, day, t) in busy_set for t in range(slot_start, slot_start + 2))

    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    wb = openpyxl.Workbook()
    thin = Side(border_style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ses, ses_label in [("morning", "Buổi Sáng"), ("afternoon", "Buổi Chiều")]:
        ws = wb.create_sheet(ses_label)

        # Header row 1: Phòng + merged day names
        hdr1 = ["Phòng"]
        for d in DAYS:
            hdr1 += [DNAMES[d], ""]
        ws.append(hdr1)

        # Header row 2: blank + slot labels
        hdr2 = [""]
        for d in DAYS:
            for s in SLOTS:
                hdr2.append(SLBL[s])
        ws.append(hdr2)

        # Merge day name cells in row 1
        for i, d in enumerate(DAYS):
            col = 2 + i * 2
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)

        # Style headers
        hdr_fill  = PatternFill("solid", fgColor="075985")
        hdr_font  = Font(color="FFFFFF", bold=True, size=9)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_idx in [1, 2]:
            for col_idx in range(1, 14):
                c = ws.cell(row_idx, col_idx)
                c.fill = hdr_fill
                c.font = hdr_font
                c.alignment = hdr_align
                c.border = border

        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 16
        ws.column_dimensions["A"].width = 18
        for col_idx in range(2, 14):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        # Data rows
        for room in all_rooms:
            row_data = [room]
            for d in DAYS:
                for s in SLOTS:
                    bk = book_map.get((ses, room, d, s))
                    if bk:
                        subj = bk.subject or bk.subject_group or ""
                        row_data.append(f"K{bk.grade} – {subj}\n{bk.teacher_name}")
                    elif is_ext_busy(ses, room, d, s):
                        row_data.append("Bận")
                    else:
                        row_data.append("")
            ws.append(row_data)
            r_idx = ws.max_row
            ws.row_dimensions[r_idx].height = 32
            for col_idx in range(1, 14):
                c = ws.cell(r_idx, col_idx)
                c.alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)
                c.border = border
                if col_idx == 1:
                    c.font = Font(bold=True, size=8)
                    c.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    c.font = Font(size=8)
                    # Color booked cells by grade
                    bk = book_map.get((ses, room, DAYS[(col_idx - 2) // 2],
                                       SLOTS[(col_idx - 2) % 2]))
                    if bk:
                        grade_colors = {10: "DBEAFE", 11: "DCFCE7", 12: "FFF7ED"}
                        fill_color = grade_colors.get(bk.grade, "F1F5F9")
                        c.fill = PatternFill("solid", fgColor=fill_color)

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="lich_phong_chi_tiet.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/register-class", methods=["POST"])
@admin_required
def admin_register_class():
    data = request.get_json(force=True)
    try:
        teacher_id    = int(data["teacher_id"])
        grade         = int(data["grade"])
        day_of_week   = int(data["day_of_week"])
        session_type  = data["session_type"]
        start_session = int(data["start_session"])
        duration      = int(data.get("duration", 2))
        location      = (data.get("location") or "").strip() or None
    except (KeyError, ValueError, TypeError):
        return jsonify(ok=False, error="Dữ liệu không hợp lệ.")

    if grade not in (10, 11, 12):
        return jsonify(ok=False, error="Khối phải là 10, 11 hoặc 12.")
    if duration not in (2, 4):
        return jsonify(ok=False, error="Số tiết phải là 2 hoặc 4.")
    if session_type not in ("morning", "afternoon"):
        return jsonify(ok=False, error="Buổi không hợp lệ.")

    end_session = start_session + duration - 1
    with engine.begin() as conn:
        teacher = conn.execute(select(teachers).where(teachers.c.id == teacher_id)).first()
        if not teacher:
            return jsonify(ok=False, error="Giáo viên không tồn tại.")
        if location:
            conflict = conn.execute(
                select(classes.c.id).where(and_(
                    classes.c.day_of_week == day_of_week,
                    classes.c.session_type == session_type,
                    classes.c.location == location,
                    classes.c.start_session <= end_session,
                    (classes.c.start_session + classes.c.duration - 1) >= start_session,
                ))
            ).first()
            if conflict:
                return jsonify(ok=False, error=f"Phòng {location} đã bị đặt trong khung giờ này.")
        result = conn.execute(
            insert(classes).values(
                teacher_id=teacher_id, grade=grade, duration=duration,
                day_of_week=day_of_week, session_type=session_type,
                start_session=start_session,
                subject=teacher.subject_group,
                location=location, max_capacity=50,
                extra_data=None, is_published=1,
            )
        )
        class_id = result.inserted_primary_key[0]
    _bump(event_type="class", grade=grade)
    return jsonify(ok=True, class_id=class_id)


@app.route("/admin/class-reg")
@admin_required
def admin_class_reg():
    with engine.connect() as conn:
        all_teachers = conn.execute(
            select(teachers).order_by(teachers.c.subject_group, teachers.c.full_name)
        ).fetchall()

        all_classes = conn.execute(
            select(classes, teachers.c.full_name.label("teacher_name"),
                   teachers.c.subject_group, teachers.c.email.label("teacher_email"))
            .join(teachers, classes.c.teacher_id == teachers.c.id)
            .order_by(teachers.c.subject_group, teachers.c.full_name, classes.c.start_session)
        ).fetchall()

        enrollment_counts = {}
        for c in all_classes:
            cnt = conn.execute(
                select(func.count()).where(enrollments.c.class_id == c.id)
            ).scalar()
            enrollment_counts[c.id] = cnt

    conflict_info = {}
    for c in all_classes:
        cls_dict = {"subject": c.subject, "day_of_week": c.day_of_week,
                    "session_type": c.session_type, "start_session": c.start_session,
                    "duration": c.duration}
        conflict_info[c.id] = _class_impact(c.grade, cls_dict)

    teacher_reg_open = get_setting("teacher_reg_open", "0") == "1"
    return render_template(
        "admin/class_reg.html",
        all_teachers=all_teachers,
        all_classes=all_classes,
        enrollment_counts=enrollment_counts,
        conflict_info=conflict_info,
        teacher_reg_open=teacher_reg_open,
        day_name=day_name,
        session_label=session_label,
    )


@app.route("/admin/class-reg/export")
@admin_required
def admin_class_reg_export():
    with engine.connect() as conn:
        all_classes = conn.execute(
            select(classes, teachers.c.full_name.label("teacher_name"),
                   teachers.c.subject_group, teachers.c.email.label("teacher_email"))
            .join(teachers, classes.c.teacher_id == teachers.c.id)
            .order_by(classes.c.grade, classes.c.day_of_week,
                      classes.c.session_type, classes.c.start_session)
        ).fetchall()
        enroll_rows = conn.execute(
            select(enrollments.c.class_id, func.count().label("cnt"))
            .group_by(enrollments.c.class_id)
        ).fetchall()
    enroll_counts = {r.class_id: r.cnt for r in enroll_rows}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Đăng ký mở lớp"
    headers = ["STT", "Họ và tên GV", "Email GV", "Tổ bộ môn", "Khối", "Thứ",
               "Buổi / Tiết", "Số tiết", "Môn học", "Địa điểm",
               "Sĩ số", "HS đăng ký", "Thời gian đăng ký"]
    ws.append(headers)
    from openpyxl.styles import PatternFill, Font, Alignment
    hdr_fill = PatternFill("solid", fgColor="0369A1")
    hdr_font = Font(color="FFFFFF", bold=True)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for idx, c in enumerate(all_classes, 1):
        buoi = "Sáng" if c.session_type == "morning" else "Chiều"
        end_ses = c.start_session + c.duration - 1
        tiet_label = f"Tiết {c.start_session}" if c.duration == 1 else f"Tiết {c.start_session}-{end_ses}"
        ws.append([
            idx,
            c.teacher_name,
            c.teacher_email or "",
            c.subject_group,
            f"Khối {c.grade}",
            day_name(c.day_of_week),
            f"{buoi} – {tiet_label}",
            c.duration,
            c.subject or "",
            c.location or "",
            c.max_capacity or "",
            enroll_counts.get(c.id, 0),
            c.created_at or "",
        ])
    for col, width in zip(range(1, 14),
                          [5, 22, 26, 14, 7, 8, 16, 7, 16, 20, 7, 10, 20]):
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="dang_ky_mo_lop.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/classes/import-rooms", methods=["POST"])
@admin_required
def admin_classes_import_rooms():
    file = request.files.get("file")
    if not file:
        flash("Chưa chọn file.")
        return redirect(url_for("admin_class_reg"))
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    if "id" not in headers or "Địa điểm" not in headers or "Sĩ số" not in headers:
        flash("File thiếu cột: cần có 'id', 'Địa điểm', 'Sĩ số'.")
        return redirect(url_for("admin_class_reg"))
    count = 0
    with engine.connect() as conn:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            try:
                class_id = int(row[headers.index("id")] or 0)
            except (ValueError, TypeError):
                continue
            location = str(row[headers.index("Địa điểm")] or "").strip() or None
            try:
                max_capacity = int(row[headers.index("Sĩ số")] or 0) or 50
            except (ValueError, TypeError):
                max_capacity = 50
            conn.execute(
                update(classes).where(classes.c.id == class_id).values(
                    location=location, max_capacity=max_capacity
                )
            )
            count += 1
        conn.commit()
    flash(f"Đã cập nhật phòng học cho {count} lớp.")
    return redirect(url_for("admin_class_reg"))


@app.route("/admin/classes/<int:class_id>", methods=["PATCH"])
@admin_required
def admin_class_update(class_id):
    data = request.get_json(force=True, silent=True) or {}
    location = (data.get("location") or "").strip() or None
    try:
        max_capacity = int(data["max_capacity"]) if data.get("max_capacity") else None
    except (ValueError, TypeError):
        max_capacity = None
    with engine.begin() as conn:
        conn.execute(
            update(classes).where(classes.c.id == class_id).values(
                location=location, max_capacity=max_capacity
            )
        )
    return jsonify(ok=True, location=location, max_capacity=max_capacity)


@app.route("/admin/class-schedule")
@admin_required
def admin_class_schedule():
    with engine.connect() as conn:
        all_classes = conn.execute(
            select(classes, teachers.c.full_name.label("teacher_name"),
                   teachers.c.subject_group)
            .join(teachers, classes.c.teacher_id == teachers.c.id)
            .order_by(classes.c.grade, teachers.c.subject_group, classes.c.start_session)
        ).fetchall()
    # Build JSON-serialisable grid for JS: {ses: {"{day}_{start}": [...]}}
    grid_json = {}
    for ses in ("morning", "afternoon"):
        grid_json[ses] = {}
        for d in range(2, 8):
            for s in (1, 3):
                grid_json[ses][f"{d}_{s}"] = []
    for c in all_classes:
        key = f"{c.day_of_week}_{c.start_session}"
        if key in grid_json.get(c.session_type, {}):
            grid_json[c.session_type][key].append({
                "id": c.id,
                "grade": c.grade,
                "subject": c.subject_group or c.subject or "",
                "teacher": c.teacher_name,
            })
    # Collect all distinct subjects (sorted) for modal columns
    all_subjects = sorted({
        c.subject_group or c.subject or "" for c in all_classes if (c.subject_group or c.subject)
    })
    return render_template(
        "admin/class_schedule.html",
        grid_json=grid_json,
        all_subjects=all_subjects,
        day_name=day_name,
    )


@app.route("/admin/class-schedule/export")
@admin_required
def admin_class_schedule_export():
    with engine.connect() as conn:
        all_classes = conn.execute(
            select(classes, teachers.c.full_name.label("teacher_name"),
                   teachers.c.subject_group)
            .join(teachers, classes.c.teacher_id == teachers.c.id)
            .order_by(classes.c.session_type, classes.c.day_of_week,
                      classes.c.start_session, classes.c.grade)
        ).fetchall()

    wb = openpyxl.Workbook()
    DAYS = [2, 3, 4, 5, 6, 7]
    DAY_NAMES_VN = {2: "Thứ 2", 3: "Thứ 3", 4: "Thứ 4", 5: "Thứ 5", 6: "Thứ 6", 7: "Thứ 7"}
    SLOTS = [(1, "Tiết 1-2"), (3, "Tiết 3-4")]
    SESSION_VN = {"morning": "Buổi Sáng", "afternoon": "Buổi Chiều"}

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="0EA5E9")
    sub_fill = PatternFill("solid", fgColor="EFF6FF")

    for ses_key in ("morning", "afternoon"):
        ws = wb.create_sheet(SESSION_VN[ses_key])
        # Header row 1: empty + day names (colspan 2 each — we'll just repeat)
        ws.append(["Tiết"] + [n for d in DAYS for n in [DAY_NAMES_VN[d], ""]])
        # Header row 2: slot label + (start_session) repeated
        ws.append([""] + ["Tiết 1-2", "Tiết 3-4"] * len(DAYS))

        # Build lookup
        grid = {}
        for c in all_classes:
            if c.session_type != ses_key:
                continue
            k = (c.day_of_week, c.start_session)
            grid.setdefault(k, []).append(c)

        for slot_start, slot_label in SLOTS:
            # Collect max items per column to know how many rows
            max_items = max(
                (len(grid.get((d, slot_start), [])) for d in DAYS),
                default=1
            )
            for i in range(max(max_items, 1)):
                row = [slot_label if i == 0 else ""]
                for d in DAYS:
                    items = grid.get((d, slot_start), [])
                    if i < len(items):
                        c = items[i]
                        row.append(f"K{c.grade} — {c.teacher_name}")
                    else:
                        row.append("")
                ws.append(row)

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions["A"].width = 12
        for col in range(2, 2 + len(DAYS)):
            ws.column_dimensions[get_column_letter(col)].width = 22

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        download_name="lich_dang_ky.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/class-reg/toggle", methods=["POST"])
@admin_required
def admin_class_reg_toggle():
    current = get_setting("teacher_reg_open", "0")
    new_val = "0" if current == "1" else "1"
    set_setting("teacher_reg_open", new_val)
    return jsonify(ok=True, open=new_val == "1")


@app.route("/admin/classes/clear-all", methods=["POST"])
@admin_required
def admin_classes_clear_all():
    """Xoá toàn bộ lớp học GV đã đăng ký mở (và đăng ký học phần liên quan)."""
    with engine.begin() as conn:
        deleted = conn.execute(text("SELECT COUNT(*) FROM classes")).scalar()
        conn.execute(text("DELETE FROM enrollments"))
        conn.execute(text("DELETE FROM classes"))
    _bump(event_type="class")
    return jsonify(ok=True, deleted=deleted)


@app.route("/admin/maintenance/toggle", methods=["POST"])
@admin_required
def admin_maintenance_toggle():
    current = get_setting("maintenance_mode", "0")
    new_val = "0" if current == "1" else "1"
    set_setting("maintenance_mode", new_val)
    return jsonify(ok=True, on=new_val == "1")


@app.route("/admin/schedule-constraint/toggle", methods=["POST"])
@admin_required
def admin_schedule_constraint_toggle():
    current = get_setting("schedule_constraint", "1")
    new_val = "0" if current == "1" else "1"
    set_setting("schedule_constraint", new_val)
    return jsonify(ok=True, on=new_val == "1")


@app.route("/admin/classes/<int:class_id>/publish", methods=["POST"])
@admin_required
def admin_class_publish(class_id):
    data = request.get_json(force=True, silent=True) or {}
    publish = 1 if data.get("publish") else 0
    with engine.begin() as conn:
        conn.execute(
            update(classes).where(classes.c.id == class_id).values(is_published=publish)
        )
    return jsonify(ok=True, is_published=bool(publish))


@app.route("/admin/classes/<int:class_id>", methods=["DELETE"])
@admin_required
def admin_class_delete(class_id):
    with engine.begin() as conn:
        conn.execute(delete(enrollments).where(enrollments.c.class_id == class_id))
        conn.execute(delete(classes).where(classes.c.id == class_id))
    return jsonify(ok=True)


# ---------------------------------------------------------------------------
# Operator routes — separate /op/ namespace, never exposes /admin/ to operators
# ---------------------------------------------------------------------------

@app.route("/op/room-detail")
@operator_required
def op_room_detail():
    with engine.connect() as conn:
        tj = _teachers_json_list(conn)
    return render_template("admin/room_grid.html",
                           teachers_json=tj,
                           is_operator_session=True,
                           api_room_grid=url_for("op_room_grid"),
                           api_register_class=url_for("op_register_class"),
                           api_class_base="/op/classes",
                           api_export_url=url_for("op_room_detail_export"))


@app.route("/op/room-grid")
@operator_required
def op_room_grid():
    return admin_room_grid.__wrapped__()


@app.route("/op/room-detail/export")
@operator_required
def op_room_detail_export():
    return admin_room_detail_export.__wrapped__()


@app.route("/op/register-class", methods=["POST"])
@operator_required
def op_register_class():
    return admin_register_class.__wrapped__()


@app.route("/op/classes/<int:class_id>", methods=["DELETE"])
@operator_required
def op_class_delete(class_id):
    return admin_class_delete.__wrapped__(class_id)


# --- Admin: Enrollment management ---

@app.route("/admin/enrollment")
@admin_required
def admin_enrollment():
    with engine.connect() as conn:
        published_classes = conn.execute(
            select(classes, teachers.c.full_name.label("teacher_name"),
                   teachers.c.subject_group)
            .join(teachers, classes.c.teacher_id == teachers.c.id)
            .where(classes.c.is_published == 1)
            .order_by(classes.c.grade, classes.c.day_of_week,
                      classes.c.session_type, classes.c.start_session)
        ).fetchall()

        enrollment_counts = {}
        for c in published_classes:
            cnt = conn.execute(
                select(func.count()).where(enrollments.c.class_id == c.id)
            ).scalar()
            enrollment_counts[c.id] = cnt

    # Collect extra_data column names across all published classes
    extra_columns = set()
    for c in published_classes:
        if c.extra_data:
            try:
                d = json.loads(c.extra_data)
                extra_columns.update(d.keys())
            except Exception:
                pass

    student_reg_open = get_setting("student_reg_open", "0") == "1"
    return render_template(
        "admin/enrollment.html",
        published_classes=published_classes,
        enrollment_counts=enrollment_counts,
        extra_columns=sorted(extra_columns),
        student_reg_open=student_reg_open,
        day_name=day_name,
        session_label=session_label,
    )


@app.route("/admin/enrollment/upload", methods=["POST"])
@admin_required
def admin_enrollment_upload():
    file = request.files.get("file")
    if not file:
        flash("Chưa chọn file.")
        return redirect(url_for("admin_enrollment"))

    wb = openpyxl.load_workbook(file)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    required = ["id", "Địa điểm", "Sĩ số"]
    for r in required:
        if r not in headers:
            flash(f"Thiếu cột: {r}")
            return redirect(url_for("admin_enrollment"))

    extra_cols = [h for h in headers if h not in required and h]
    count = 0
    with engine.connect() as conn:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            try:
                class_id = int(row[headers.index("id")] or 0)
            except (ValueError, TypeError):
                continue
            location = str(row[headers.index("Địa điểm")] or "").strip()
            try:
                max_capacity = int(row[headers.index("Sĩ số")] or 0) or 50
            except (ValueError, TypeError):
                max_capacity = 50

            extra = {}
            for col in extra_cols:
                val = row[headers.index(col)]
                if val is not None:
                    extra[col] = str(val)

            extra_data_json = json.dumps(extra, ensure_ascii=False) if extra else None

            conn.execute(
                update(classes).where(classes.c.id == class_id).values(
                    location=location,
                    max_capacity=max_capacity,
                    extra_data=extra_data_json,
                    is_published=1,
                )
            )
            count += 1
        conn.commit()

    flash(f"Đã cập nhật {count} lớp và mở đăng ký cho học sinh.")
    return redirect(url_for("admin_enrollment"))


@app.route("/admin/enrollment/add", methods=["POST"])
@admin_required
def admin_enrollment_add():
    data = request.get_json(force=True)
    try:
        teacher_id = int(data["teacher_id"])
        grade = int(data["grade"])
        day_of_week = int(data["day_of_week"])
        session_type = data["session_type"]
        start_session = int(data["start_session"])
        duration = int(data["duration"])
        subject = (data.get("subject") or "").strip() or None
        location = (data.get("location") or "").strip() or None
        max_capacity = int(data["max_capacity"]) if data.get("max_capacity") else 50
        extra_data_json = data.get("extra_data_json") or None
    except (KeyError, ValueError, TypeError) as e:
        return jsonify(ok=False, error=f"Dữ liệu không hợp lệ: {e}")

    with engine.connect() as conn:
        result = conn.execute(
            insert(classes).values(
                teacher_id=teacher_id,
                grade=grade,
                duration=duration,
                day_of_week=day_of_week,
                session_type=session_type,
                start_session=start_session,
                subject=subject,
                location=location,
                max_capacity=max_capacity,
                extra_data=extra_data_json,
                is_published=1,
                created_at=now_vn(),
            )
        )
        conn.commit()
        class_id = result.inserted_primary_key[0]

    return jsonify(ok=True, class_id=class_id)


@app.route("/admin/enrollment/export")
@admin_required
def admin_enrollment_export():
    with engine.connect() as conn:
        published_classes = conn.execute(
            select(classes, teachers.c.full_name.label("teacher_name"),
                   teachers.c.subject_group, teachers.c.email.label("teacher_email"))
            .join(teachers, classes.c.teacher_id == teachers.c.id)
            .where(classes.c.is_published == 1)
            .order_by(classes.c.grade, classes.c.day_of_week,
                      classes.c.session_type, classes.c.start_session)
        ).fetchall()

    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    summary_rows = []

    with engine.connect() as conn:
        for cls in published_classes:
            sheet_name = f"Lớp {cls.id}: {cls.subject or 'N/A'} K{cls.grade}"
            # Excel sheet names max 31 chars
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["STT", "Họ tên", "Lớp", "Khối"])

            enrolled_students = conn.execute(
                select(students, enrollments.c.enrolled_at)
                .join(enrollments, students.c.id == enrollments.c.student_id)
                .where(enrollments.c.class_id == cls.id)
                .order_by(students.c.class_name, students.c.full_name)
            ).fetchall()

            for i, s in enumerate(enrolled_students, 1):
                ws.append([i, s.full_name, s.class_name, s.grade])

            enrolled_count = len(enrolled_students)
            buoi = "Sáng" if cls.session_type == "morning" else "Chiều"
            summary_rows.append([
                f"Lớp {cls.id}",
                cls.teacher_name,
                cls.subject or "",
                cls.grade,
                day_name(cls.day_of_week),
                buoi,
                cls.start_session,
                cls.location or "",
                enrolled_count,
                cls.max_capacity or "",
            ])

    # Summary sheet
    ws_sum = wb.create_sheet(title="Tổng hợp", index=0)
    ws_sum.append(["Lớp", "GV", "Môn", "Khối", "Thứ", "Buổi", "Tiết",
                   "Địa điểm", "Sĩ số đã đăng ký", "Sĩ số tối đa"])
    for row in summary_rows:
        ws_sum.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="danh_sach_dang_ky.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/enrollment/toggle", methods=["POST"])
@admin_required
def admin_enrollment_toggle():
    current = get_setting("student_reg_open", "0")
    new_val = "0" if current == "1" else "1"
    set_setting("student_reg_open", new_val)
    return jsonify(ok=True, open=new_val == "1")


@app.route("/admin/enrollment/<int:class_id>/students")
@admin_required
def admin_enrollment_students(class_id):
    with engine.connect() as conn:
        enrolled = conn.execute(
            select(students, enrollments.c.enrolled_at)
            .join(enrollments, students.c.id == enrollments.c.student_id)
            .where(enrollments.c.class_id == class_id)
            .order_by(students.c.class_name, students.c.full_name)
        ).fetchall()

    return jsonify(ok=True, students=[
        {
            "id": s.id,
            "full_name": s.full_name,
            "cccd": s.cccd,
            "class_name": s.class_name,
            "grade": s.grade,
            "enrolled_at": s.enrolled_at,
            "last_seen_at": s.last_seen_at if hasattr(s, "last_seen_at") else None,
        }
        for s in enrolled
    ])

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    app.run(debug=True, port=5051, threaded=True)
